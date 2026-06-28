"""Leave dashboard analytics and reporting."""

from __future__ import annotations

import calendar
import csv
import io
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from django.utils import timezone

from apps.accounts.models import User
from apps.organizations.models import Department, Organization

from .models import Holiday, LeaveRequest, LeaveType
from .services import leave_team_for


@dataclass
class LeaveFilters:
    date_from: date
    date_to: date
    department: str = ""
    branch: str = ""
    employee_id: str = ""
    leave_type_id: str = ""
    status: str = ""
    search: str = ""

    @classmethod
    def from_request(cls, request):
        today = timezone.localdate()
        date_from = _parse_date(request.GET.get("from"), today.replace(day=1))
        date_to = _parse_date(request.GET.get("to"), today)
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        return cls(
            date_from=date_from,
            date_to=date_to,
            department=(request.GET.get("department") or "").strip(),
            branch=(request.GET.get("branch") or "").strip(),
            employee_id=(request.GET.get("employee") or "").strip(),
            leave_type_id=(request.GET.get("leave_type") or "").strip(),
            status=(request.GET.get("status") or "").strip(),
            search=(request.GET.get("q") or "").strip(),
        )


def _parse_date(raw, default: date) -> date:
    if raw:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            pass
    return default


def filtered_requests(viewer: User, filters: LeaveFilters):
    team = leave_team_for(viewer)
    if filters.employee_id:
        team = team.filter(pk=filters.employee_id)
    if filters.department:
        team = team.filter(department_id=filters.department)
    if filters.branch:
        team = team.filter(work_location__iexact=filters.branch)
    team_ids = list(team.values_list("pk", flat=True))

    qs = (
        LeaveRequest.objects.filter(
            user_id__in=team_ids,
            start_date__lte=filters.date_to,
            end_date__gte=filters.date_from,
        )
        .select_related("user", "user__department", "leave_type", "reviewed_by")
        .prefetch_related("approvals")
    )

    if filters.leave_type_id:
        qs = qs.filter(leave_type_id=filters.leave_type_id)
    if filters.status:
        qs = qs.filter(status=filters.status)
    if filters.search:
        qs = qs.filter(
            Q(user__first_name__icontains=filters.search)
            | Q(user__last_name__icontains=filters.search)
            | Q(user__employee_id__icontains=filters.search)
            | Q(reason__icontains=filters.search)
        )
    return qs.order_by("-applied_at")


def build_summary(viewer: User, filters: LeaveFilters) -> dict:
    today = timezone.localdate()
    month_start = today.replace(day=1)
    qs = filtered_requests(viewer, filters)
    all_org_qs = qs

    pending = qs.filter(status=LeaveRequest.Status.PENDING).count()
    approved = qs.filter(status=LeaveRequest.Status.APPROVED).count()
    rejected = qs.filter(status=LeaveRequest.Status.REJECTED).count()
    total = qs.count()

    on_leave_today = qs.filter(
        status=LeaveRequest.Status.APPROVED,
        start_date__lte=today,
        end_date__gte=today,
    ).count()

    sick = LeaveType.objects.filter(code="sick-leave").first()
    casual = LeaveType.objects.filter(code="casual-leave").first()
    sick_month = 0
    casual_month = 0
    if sick:
        sick_month = all_org_qs.filter(
            leave_type=sick,
            applied_at__date__gte=month_start,
            status__in=[LeaveRequest.Status.APPROVED, LeaveRequest.Status.PENDING],
        ).count()
    if casual:
        casual_month = all_org_qs.filter(
            leave_type=casual,
            applied_at__date__gte=month_start,
            status__in=[LeaveRequest.Status.APPROVED, LeaveRequest.Status.PENDING],
        ).count()

    avg_days = (
        qs.filter(status=LeaveRequest.Status.APPROVED).aggregate(a=Avg("total_days"))["a"] or Decimal("0")
    )

    prev_end = filters.date_from - timedelta(days=1)
    prev_start = prev_end - (filters.date_to - filters.date_from)
    prev_total = LeaveRequest.objects.filter(
        user__organization=viewer.organization,
        applied_at__date__gte=prev_start,
        applied_at__date__lte=prev_end,
    ).count()
    trend = round(((total - prev_total) / prev_total) * 100, 1) if prev_total else (100 if total else 0)

    return {
        "total_requests": total,
        "pending": pending,
        "approved": approved,
        "rejected": rejected,
        "on_leave_today": on_leave_today,
        "sick_month": sick_month,
        "casual_month": casual_month,
        "avg_usage": f"{avg_days:.1f}d" if avg_days else "0d",
        "trend": trend,
    }


def build_charts(viewer: User, filters: LeaveFilters) -> dict:
    qs = filtered_requests(viewer, filters)
    monthly: dict[str, int] = defaultdict(int)
    dept: dict[str, int] = defaultdict(int)
    type_dist: dict[str, int] = defaultdict(int)

    for req in qs:
        key = req.applied_at.strftime("%Y-%m")
        monthly[key] += 1
        dept[req.user.department_name or "Unassigned"] += 1
        type_dist[req.leave_type.name] += 1

    months = sorted(monthly.keys())[-12:]
    return {
        "monthly": {"labels": months, "values": [monthly[m] for m in months]},
        "department": {
            "labels": list(dept.keys())[:10],
            "values": [dept[k] for k in list(dept.keys())[:10]],
        },
        "types": {
            "labels": list(type_dist.keys()),
            "values": list(type_dist.values()),
        },
    }


def build_insights(viewer: User, filters: LeaveFilters) -> list[dict]:
    qs = filtered_requests(viewer, filters)
    insights = []

    by_user: dict = defaultdict(int)
    for req in qs.filter(status=LeaveRequest.Status.APPROVED):
        by_user[req.user_id] += 1
    if by_user:
        top_uid = max(by_user, key=by_user.get)
        top_user = User.objects.filter(pk=top_uid).first()
        if top_user and by_user[top_uid] >= 3:
            insights.append(
                {
                    "icon": "alert-triangle",
                    "title": "High leave frequency",
                    "body": f"{top_user.choice_label} has {by_user[top_uid]} approved leaves in this period.",
                }
            )

    pending = qs.filter(status=LeaveRequest.Status.PENDING).count()
    if pending:
        insights.append(
            {
                "icon": "clock",
                "title": "Pending approvals",
                "body": f"{pending} request(s) awaiting manager or HR approval.",
            }
        )

    peak = (
        qs.filter(status=LeaveRequest.Status.APPROVED)
        .values("start_date__month")
        .annotate(c=Count("id"))
        .order_by("-c")
        .first()
    )
    if peak and peak["c"] > 2:
        month_name = calendar.month_name[peak["start_date__month"]]
        insights.append(
            {
                "icon": "trending-up",
                "title": "Peak leave period",
                "body": f"Highest activity in {month_name} ({peak['c']} requests).",
            }
        )

    insights.append(
        {
            "icon": "sparkles",
            "title": "Leave prediction",
            "body": "Enable accrual policies to forecast balance usage next quarter.",
        }
    )
    return insights[:6]


def build_calendar_events(viewer: User, year: int, month: int, org: Organization) -> list[dict]:
    _, last = calendar.monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, last)
    team = leave_team_for(viewer)
    team_ids = list(team.values_list("pk", flat=True))

    events = []
    for req in LeaveRequest.objects.filter(
        user_id__in=team_ids,
        start_date__lte=end,
        end_date__gte=start,
        status__in=[LeaveRequest.Status.PENDING, LeaveRequest.Status.APPROVED],
    ).select_related("user", "leave_type"):
        color = "#34d399" if req.status == LeaveRequest.Status.APPROVED else "#fbbf24"
        if req.status == LeaveRequest.Status.REJECTED:
            color = "#fb7185"
        events.append(
            {
                "title": f"{req.user.first_name} · {req.leave_type.code}",
                "start": req.start_date.isoformat(),
                "end": req.end_date.isoformat(),
                "color": color,
                "status": req.status,
                "id": str(req.pk),
            }
        )

    for hol in Holiday.objects.filter(organization=org, date__gte=start, date__lte=end):
        events.append(
            {
                "title": hol.name,
                "start": hol.date.isoformat(),
                "end": hol.date.isoformat(),
                "color": "#a78bfa",
                "status": "HOLIDAY",
                "id": str(hol.pk),
            }
        )
    return events


def table_rows(qs) -> list[dict]:
    rows = []
    for req in qs:
        rows.append(
            {
                "id": req.pk,
                "employee_id": req.user.employee_id or "—",
                "employee_name": req.user.choice_label,
                "department": req.user.department_name or "—",
                "leave_type": req.leave_type.name,
                "start_date": req.start_date,
                "end_date": req.end_date,
                "days": req.total_days,
                "status": req.status,
                "status_display": req.get_status_display(),
                "stage": req.current_stage_label,
                "applied_at": req.applied_at,
                "approved_by": req.reviewed_by.choice_label if req.reviewed_by else "—",
                "reason": req.reason[:80] + ("…" if len(req.reason) > 80 else ""),
                "user_id": req.user_id,
                "can_approve": req.status == LeaveRequest.Status.PENDING,
            }
        )
    return rows


def filter_options(viewer: User, org: Organization) -> dict:
    team = leave_team_for(viewer)
    return {
        "departments": Department.objects.filter(organization=org, is_active=True).order_by("name"),
        "branches": list(
            team.exclude(work_location="")
            .values_list("work_location", flat=True)
            .distinct()
            .order_by("work_location")
        ),
        "employees": list(team.order_by("first_name", "last_name")),
        "leave_types": list(LeaveType.objects.filter(organization=org, is_active=True)),
        "status_choices": LeaveRequest.Status.choices,
        "holidays": list(Holiday.objects.filter(organization=org).order_by("date")),
    }


def export_csv(rows: list[dict]) -> HttpResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(
        [
            "Employee ID",
            "Name",
            "Department",
            "Leave Type",
            "Start",
            "End",
            "Days",
            "Status",
            "Applied",
            "Approved By",
            "Reason",
        ]
    )
    for r in rows:
        w.writerow(
            [
                r["employee_id"],
                r["employee_name"],
                r["department"],
                r["leave_type"],
                r["start_date"],
                r["end_date"],
                r["days"],
                r["status_display"],
                r["applied_at"],
                r["approved_by"],
                r["reason"],
            ]
        )
    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="leave_report.csv"'
    return resp


_REPORT_HEADER = [
    "Employee ID",
    "Name",
    "Department",
    "Leave Type",
    "Start",
    "End",
    "Days",
    "Status",
    "Applied",
    "Approved By",
    "Reason",
]


def export_xlsx(rows: list[dict]) -> HttpResponse:
    """Excel (.xlsx) version of the leave report export."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave report"
    ws.append(_REPORT_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append(
            [
                r["employee_id"],
                r["employee_name"],
                r["department"],
                r["leave_type"],
                r["start_date"],
                r["end_date"],
                float(r["days"]),
                r["status_display"],
                r["applied_at"].replace(tzinfo=None) if r["applied_at"] else None,
                r["approved_by"],
                r["reason"],
            ]
        )
    out = io.BytesIO()
    wb.save(out)
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="leave_report.xlsx"'
    return resp


def balance_report_rows(viewer: User) -> list[dict]:
    """Per-user × leave-type balance rows for the viewer's leave team."""
    from django.utils import timezone

    from .models import LeaveBalance

    year = timezone.localdate().year
    team_ids = list(leave_team_for(viewer).values_list("pk", flat=True))
    rows = []
    qs = (
        LeaveBalance.objects.filter(user_id__in=team_ids, year=year)
        .select_related("user", "user__department", "leave_type")
        .order_by("user__first_name", "user__last_name", "leave_type__name")
    )
    for bal in qs:
        rows.append(
            {
                "employee_id": bal.user.employee_id or "—",
                "employee_name": bal.user.choice_label,
                "department": bal.user.department_name or "—",
                "leave_type": bal.leave_type.name,
                "year": bal.year,
                "allocated": bal.allocated,
                "carried_forward": bal.carried_forward,
                "adjusted": bal.adjusted,
                "used": bal.used,
                "remaining": bal.remaining,
            }
        )
    return rows


_BALANCE_HEADER = [
    "Employee ID",
    "Name",
    "Department",
    "Leave Type",
    "Year",
    "Allocated",
    "Carried Forward",
    "Adjusted",
    "Used",
    "Remaining",
]


def export_balance_csv(rows: list[dict]) -> HttpResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_BALANCE_HEADER)
    for r in rows:
        w.writerow([r[k] for k in (
            "employee_id", "employee_name", "department", "leave_type", "year",
            "allocated", "carried_forward", "adjusted", "used", "remaining",
        )])
    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="leave_balances.csv"'
    return resp


def export_balance_xlsx(rows: list[dict]) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave balances"
    ws.append(_BALANCE_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([
            r["employee_id"], r["employee_name"], r["department"], r["leave_type"],
            r["year"], float(r["allocated"]), float(r["carried_forward"]),
            float(r["adjusted"]), float(r["used"]), float(r["remaining"]),
        ])
    out = io.BytesIO()
    wb.save(out)
    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="leave_balances.xlsx"'
    return resp
