import calendar
import json
from decimal import Decimal

from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.generic import TemplateView

from apps.accounts.models import User
from apps.dashboard.mixins import AdminOrHRRequiredMixin, AdminRequiredMixin, OrganizationRequiredMixin
from apps.organizations.module_utils import ensure_module, plan_includes_module
from apps.organizations.financial_year import payroll_month_choices, get_current_financial_year
from apps.organizations.fy_utils import get_selected_fy
from apps.subscriptions.mixins import PlanFeatureRequiredMixin

from .action_mixins import PayrollActionMixin


def _get_period_months(org, fy: dict | None = None):
    """Return payroll month choices limited to the selected FY period."""
    try:
        from datetime import date as _date
        today = timezone.localdate()
        if fy:
            # Walk every month inside the FY up to today
            months = []
            d = fy["date_from"].replace(day=1)
            end = min(fy["date_to"], today)
            while d <= end:
                months.append({
                    "month": d.month,
                    "year": d.year,
                    "label": f"{calendar.month_name[d.month]} {d.year}",
                    "short": f"{calendar.month_abbr[d.month]} {d.year}",
                })
                d = d.replace(month=d.month + 1) if d.month < 12 else d.replace(year=d.year + 1, month=1)
            return list(reversed(months))  # most recent first
        return payroll_month_choices(org, n_years=2)
    except Exception:
        today = timezone.localdate()
        return [
            {"month": i, "year": today.year, "label": f"{calendar.month_name[i]} {today.year}"}
            for i in range(1, 13)
        ]


def _get_period_years(months):
    """Deduplicated years from period_months, most recent first."""
    seen, result = set(), []
    for m in months:
        y = m["year"]
        if y not in seen:
            seen.add(y)
            result.append(y)
    return result


from .analytics import (
    PayrollFilters,
    approval_steps,
    build_charts,
    build_insights,
    build_summary,
    export_csv,
    filter_options,
    filtered_payslips,
    get_current_run,
    pending_reimbursements,
    recent_revisions,
    recent_runs,
    salary_components_panel,
    structures_panel,
    table_rows,
)
from .forms import ReimbursementForm, SalaryRevisionForm
from .models import EmployeeSalary, Payslip, PayrollRun, Reimbursement, SalaryComponent, SalaryRevision
from .services import (
    approve_payroll_run,
    build_salary_structure_rows,
    ensure_payroll_setup,
    generate_payslip_numbers,
    get_or_create_payroll_run,
    lock_payroll_run,
    mark_payroll_paid,
    process_payroll_run,
)


class PayrollManagementView(OrganizationRequiredMixin, TemplateView):
    template_name = "payroll/payroll_management.html"
    paginate_by = 20

    def _redirect_back(self, request):
        y = request.POST.get("year") or request.GET.get("year")
        m = request.POST.get("month") or request.GET.get("month")
        url = reverse("payroll:management")
        if y and m:
            return redirect(f"{url}?year={y}&month={m}")
        return redirect(url)

    def dispatch(self, request, *args, **kwargs):
        org = getattr(request.user, "organization", None)
        if org:
            active, synced = ensure_module(org, "payroll", getattr(request.user, "role", None))
            if synced:
                messages.info(request, "Payroll has been enabled for your organization.")
            elif not active:
                if plan_includes_module(org, "payroll"):
                    messages.warning(
                        request,
                        "Payroll is disabled for your organization. "
                        "An admin can re-enable it under Settings → HR modules.",
                    )
                else:
                    messages.warning(request, "Payroll is not included in your subscription plan.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "csv":
            fy = get_selected_fy(request)
            filters = PayrollFilters.from_request(request, fy=fy)
            rows = table_rows(filtered_payslips(request.user, filters))
            return export_csv(rows)
        if request.GET.get("payslip"):
            return self._payslip_preview(request)
        return super().get(request, *args, **kwargs)

    def _payslip_preview(self, request):
        slip = get_object_or_404(
            Payslip.objects.select_related("user", "payroll_run").prefetch_related("lines"),
            pk=request.GET.get("payslip"),
            user__organization=request.user.organization,
        )
        # Only ADMIN/HR (finance) may view other people's payslips. Managers and
        # employees are restricted to their own — a manager has no payroll access
        # to their team's payslips.
        if request.user.role not in (User.Role.ADMIN, User.Role.HR) and slip.user_id != request.user.pk:
            messages.error(request, "You can only view your own payslip.")
            return redirect("payroll:management")
        # Track distribution: opening a payslip counts as a download.
        from django.db.models import F

        Payslip.objects.filter(pk=slip.pk).update(
            download_count=F("download_count") + 1, last_downloaded_at=timezone.now()
        )
        from .pdf import _payslip_format_for, payslip_common_context
        from .payslip_formats import preview_template_for

        ctx = payslip_common_context(slip)
        ctx["preview_template"] = preview_template_for(
            _payslip_format_for(request.user.organization)
        )
        return render(request, "payroll/payslip_preview.html", ctx)

    def post(self, request, *args, **kwargs):
        org = request.user.organization
        action = request.POST.get("action")
        user = request.user

        if action == "run_payroll" and user.role in (User.Role.ADMIN, User.Role.HR):
            return self._run_payroll(request, org)
        if action == "approve_payroll" and user.role == User.Role.ADMIN:
            return self._approve_payroll(request, org)
        if action == "mark_paid" and user.role == User.Role.ADMIN:
            return self._mark_paid(request, org)
        if action == "lock_payroll" and user.role == User.Role.ADMIN:
            return self._lock_payroll(request, org)
        if action == "generate_payslips" and user.role in (User.Role.ADMIN, User.Role.HR):
            return self._generate_payslips(request, org)
        if action == "add_reimbursement":
            return self._add_reimbursement(request)
        if action == "approve_reimbursement" and user.role in (User.Role.ADMIN, User.Role.HR):
            return self._approve_reimbursement(request)
        if action == "reject_reimbursement" and user.role in (User.Role.ADMIN, User.Role.HR):
            return self._reject_reimbursement(request)
        if action == "salary_revision" and user.role == User.Role.ADMIN:
            return self._salary_revision(request)
        if action == "save_payroll_policy" and user.role == User.Role.ADMIN:
            return self._save_payroll_policy(request, org)

        messages.error(request, "Invalid action.")
        return self._redirect_back(request)

    def _save_payroll_policy(self, request, org):
        from apps.organizations.models import Organization

        policy = request.POST.get("payroll_lop_policy")
        if policy in Organization.PayrollLopPolicy.values:
            org.payroll_lop_policy = policy
            org.save(update_fields=["payroll_lop_policy", "updated_at"])
            messages.success(request, "Payroll pay policy updated.")
        else:
            messages.error(request, "Invalid payroll policy.")
        return self._redirect_back(request)

    def _filters_from_request(self, request) -> PayrollFilters:
        if request.method == "POST" and (request.POST.get("year") or request.POST.get("month")):
            today = timezone.localdate()
            year = int(request.POST.get("year") or today.year)
            month = int(request.POST.get("month") or today.month)
            return PayrollFilters(year=year, month=month)
        return PayrollFilters.from_request(request)

    def _run_payroll(self, request, org):
        from .models import PayrollAuditLog
        from .services import record_payroll_action

        filters = self._filters_from_request(request)
        run = get_or_create_payroll_run(org, filters.year, filters.month)
        msg = process_payroll_run(run, request.user)
        record_payroll_action(
            org, request.user, PayrollAuditLog.Action.PROCESSED,
            f"Payroll processed for {run.period_label}",
            period=run.period_label, request=request, employees=run.employee_count,
        )
        messages.success(request, msg)
        return self._redirect_back(request)

    def _approve_payroll(self, request, org):
        run = self._get_run(request, org)
        if run is None:
            return self._run_required_error(request)
        msg = approve_payroll_run(run, request.user, request.POST.get("comment", ""))
        messages.success(request, msg)
        return self._redirect_back(request)

    def _mark_paid(self, request, org):
        run = self._get_run(request, org)
        if run is None:
            return self._run_required_error(request)
        msg = mark_payroll_paid(run)
        messages.success(request, msg)
        return self._redirect_back(request)

    def _lock_payroll(self, request, org):
        run = self._get_run(request, org)
        if run is None:
            return self._run_required_error(request)
        msg = lock_payroll_run(run)
        messages.warning(request, msg)
        return self._redirect_back(request)

    def _generate_payslips(self, request, org):
        run = self._get_run(request, org)
        if run is None:
            return self._run_required_error(request)
        count = generate_payslip_numbers(run)
        messages.success(request, f"Generated {count} payslip(s).")
        return self._redirect_back(request)

    def _get_run(self, request, org):
        filters = self._filters_from_request(request)
        return PayrollRun.objects.filter(
            organization=org,
            year=filters.year,
            month=filters.month,
        ).first()

    def _run_required_error(self, request):
        messages.error(
            request,
            "No payroll run exists for this period yet. Use 'Calculate' to create it first.",
        )
        return self._redirect_back(request)

    def _add_reimbursement(self, request):
        if request.user.role == User.Role.ADMIN:
            messages.error(request, "Admins approve reimbursements; submit as HR or employee.")
            return self._redirect_back(request)
        form = ReimbursementForm(request.POST, request.FILES)
        if form.is_valid():
            reimb = form.save(commit=False)
            reimb.user = request.user
            reimb.save()
            messages.success(request, "Reimbursement claim submitted.")
        else:
            messages.error(request, "Invalid reimbursement form.")
        return self._redirect_back(request)

    def _approve_reimbursement(self, request):
        reimb = get_object_or_404(Reimbursement, pk=request.POST.get("reimbursement_id"))
        reimb.status = Reimbursement.Status.APPROVED
        reimb.reviewed_by = request.user
        reimb.reviewed_at = timezone.now()
        reimb.save()
        messages.success(request, "Reimbursement approved.")
        return self._redirect_back(request)

    def _reject_reimbursement(self, request):
        reimb = get_object_or_404(Reimbursement, pk=request.POST.get("reimbursement_id"))
        reimb.status = Reimbursement.Status.REJECTED
        reimb.reviewed_by = request.user
        reimb.reviewed_at = timezone.now()
        reimb.save()
        messages.warning(request, "Reimbursement rejected.")
        return self._redirect_back(request)

    def _salary_revision(self, request):
        target = get_object_or_404(User, pk=request.POST.get("user_id"), organization=request.user.organization)
        form = SalaryRevisionForm(request.POST)
        if form.is_valid():
            from .services import get_active_salary

            profile = get_active_salary(target)
            rev = form.save(commit=False)
            rev.user = target
            rev.previous_ctc = profile.monthly_ctc
            rev.status = SalaryRevision.Status.APPROVED
            rev.approved_by = request.user
            rev.save()
            profile.monthly_ctc = rev.new_ctc
            profile.save(update_fields=["monthly_ctc"])
            messages.success(request, f"Salary revised for {target.choice_label}.")
        else:
            messages.error(request, "Invalid salary revision.")
        return self._redirect_back(request)

    def get_context_data(self, **kwargs):
        from .compliance import can_view_compliance

        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = user.organization
        ensure_payroll_setup(org)

        fy = get_selected_fy(self.request)
        filters = PayrollFilters.from_request(self.request, fy=fy)
        run = get_current_run(org, filters)
        qs = filtered_payslips(user, filters)
        paginator = Paginator(table_rows(qs), self.paginate_by)
        page = paginator.get_page(self.request.GET.get("page") or 1)

        query = self.request.GET.copy()
        query.pop("page", None)
        query.pop("export", None)
        query.pop("payslip", None)

        charts = build_charts(user, filters, fy=fy)
        is_admin = user.role == User.Role.ADMIN
        is_finance = user.role in (User.Role.ADMIN, User.Role.HR)

        context.update(
            {
                "organization": org,
                "filters": filters,
                "filters_get": self.request.GET,
                "filter_query": query.urlencode(),
                "summary": build_summary(user, filters),
                "charts_json": json.dumps(charts),
                "insights": build_insights(user, filters),
                "filter_options": filter_options(user, org),
                "table_rows": page.object_list,
                "page_obj": page,
                "payroll_run": run,
                "approval_steps": approval_steps(run),
                "salary_components": salary_components_panel(org),
                "salary_structures": structures_panel(org),
                "pending_reimbursements": pending_reimbursements(user),
                "recent_revisions": recent_revisions(org),
                "reimbursement_form": ReimbursementForm(),
                "is_admin": is_admin,
                "can_access_compliance": can_view_compliance(user),
                "is_finance": is_finance,
                "can_process": is_finance,
                "can_employee_claim": user.role in (User.Role.HR, User.Role.EMPLOYEE),
                "today": timezone.localdate(),
                "month_choices": [(i, calendar.month_name[i]) for i in range(1, 13)],
                "period_months": (_pm := _get_period_months(org, fy=fy)),
                "period_years": _get_period_years(_pm),
                "lop_policy": getattr(org, "payroll_lop_policy", ""),
                "lop_policy_choices": org.PayrollLopPolicy.choices,
                "selected_fy": fy,
                "run_history": [
                    {
                        "year": r.year,
                        "month": r.month,
                        "period_label": r.period_label,
                        "employee_count": r.employee_count,
                        # Expense totals shown positive (DB signs preserved).
                        "gross": abs(r.total_gross),
                        "deductions": abs(r.total_deductions),
                        "net": abs(r.total_net),
                        "status_display": r.get_status_display(),
                    }
                    for r in recent_runs(org, fy=fy)
                ]
                if is_finance
                else [],
            }
        )
        return context


class SalaryStructuresView(OrganizationRequiredMixin, TemplateView):
    """Payroll → Employee Salary Structures: per-employee salary breakdown."""

    template_name = "payroll/salary_structures.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role not in (
            User.Role.ADMIN,
            User.Role.HR,
        ):
            messages.error(request, "Only Admin and HR can manage salary structures.")
            return redirect("payroll:management")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        rows = build_salary_structure_rows(user)
        total_gross = sum((r["gross"] for r in rows), Decimal("0"))
        total_net = sum((r["net"] for r in rows), Decimal("0"))
        total_ded = sum((r["deductions"] for r in rows), Decimal("0"))
        context.update(
            {
                "organization": user.organization,
                "rows": rows,
                "summary": {
                    "count": len(rows),
                    "gross": total_gross,
                    "deductions": total_ded,
                    "net": total_net,
                },
                "is_admin": user.role == User.Role.ADMIN,
            }
        )
        return context


class BulkSalaryStructureView(OrganizationRequiredMixin, TemplateView):
    """Edit every employee's salary components in one grid (Admin/HR)."""

    template_name = "payroll/salary_structures_bulk.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.role not in (
            User.Role.ADMIN,
            User.Role.HR,
        ):
            messages.error(request, "Only Admin and HR can manage salary structures.")
            return redirect("payroll:management")
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        from .salary_grid import save_bulk_salary

        count = save_bulk_salary(request.user, request.POST)
        messages.success(request, f"Salary structures updated for {count} employee(s).")
        return redirect("payroll:salary_structures_bulk")

    def get_context_data(self, **kwargs):
        from .salary_grid import bulk_salary_grid

        context = super().get_context_data(**kwargs)
        context["organization"] = self.request.user.organization
        context.update(bulk_salary_grid(self.request.user))
        return context


class EmployeeFinancialsView(OrganizationRequiredMixin, TemplateView):
    """Dedicated salary/financial profile for one employee (CTC, breakdown,
    bank + statutory details). Not the full HR profile."""

    template_name = "payroll/employee_financials.html"
    FINANCIAL_FIELDS = [
        "bank_name", "bank_account_holder", "bank_account_number", "ifsc_code",
        "pan_number", "aadhaar_number", "pf_account_number", "uan_number", "esi_number",
    ]

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.user.role not in (User.Role.ADMIN, User.Role.HR):
                messages.error(request, "Only Admin and HR can manage employee financials.")
                return redirect("payroll:salary_structures")
            from .services import payroll_team_for

            self.employee = get_object_or_404(payroll_team_for(request.user), pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        from django.utils.dateparse import parse_date

        from .models import EmployeeSalaryComponent as ESC
        from .services import get_active_salary, seed_employee_components

        emp = self.employee
        salary = get_active_salary(emp)

        # ── Salary (CTC / type / effective) ──
        try:
            new_ctc = Decimal(request.POST.get("monthly_ctc") or "0")
        except Exception:
            new_ctc = None
        effective = parse_date(request.POST.get("effective_from") or "") or timezone.localdate()
        salary_type = request.POST.get("salary_type")

        if new_ctc is not None and new_ctc > 0:
            if new_ctc != salary.monthly_ctc:
                from apps.dashboard.notification_service import send_notification

                rev = SalaryRevision.objects.create(
                    user=emp,
                    previous_ctc=salary.monthly_ctc,
                    new_ctc=new_ctc,
                    effective_date=effective,
                    reason=(request.POST.get("revision_reason") or "").strip(),
                    status=SalaryRevision.Status.APPROVED,
                    approved_by=request.user,
                )
                send_notification(
                    emp,
                    source_key=f"salary_revised_{rev.pk}",
                    title="Salary revised",
                    message=f"Your salary has been revised, effective {effective:%d %b %Y}.",
                    url=reverse("payroll:dashboard"),
                    icon="trending-up",
                )
            salary.monthly_ctc = new_ctc
        if salary_type in EmployeeSalary.SalaryType.values:
            salary.salary_type = salary_type
        salary.effective_from = effective
        salary.save()

        # ── Editable salary components (per employee) ──
        seed_employee_components(salary)
        for comp in salary.components.all():
            mode = request.POST.get(f"comp_{comp.pk}_mode")
            value = request.POST.get(f"comp_{comp.pk}_value")
            changed = False
            if mode in ESC.Mode.values and mode != comp.mode:
                comp.mode = mode
                changed = True
            if value is not None:
                try:
                    v = Decimal(value)
                    if v != comp.value:
                        comp.value = v
                        changed = True
                except Exception:
                    pass
            if changed:
                comp.save(update_fields=["mode", "value"])

        # ── Bank + statutory details on the user ──
        for field in self.FINANCIAL_FIELDS:
            setattr(emp, field, (request.POST.get(field) or "").strip())
        emp.save(update_fields=self.FINANCIAL_FIELDS)

        messages.success(request, f"Saved financial details for {emp.display_name}.")
        return redirect("payroll:employee_financials", pk=emp.pk)

    def get_context_data(self, **kwargs):
        from .services import compute_employee_breakdown, get_active_salary, seed_employee_components

        context = super().get_context_data(**kwargs)
        emp = self.employee
        salary = get_active_salary(emp)
        seed_employee_components(salary)
        components = list(salary.components.all().order_by("sort_order"))
        org_cat_map = {
            sc.code: sc.category
            for sc in SalaryComponent.objects.filter(organization=emp.organization, is_active=True)
        }
        comp_json = [
            {
                "id": str(c.pk), "code": c.code, "label": c.label,
                "kind": c.kind, "mode": c.mode, "value": float(c.value),
                "category": org_cat_map.get(c.code, ""),
            }
            for c in components
        ]
        context.update(
            {
                "organization": self.request.user.organization,
                "employee": emp,
                "salary": salary,
                "breakdown": compute_employee_breakdown(salary),
                "components_json": comp_json,
                "salary_type_choices": EmployeeSalary.SalaryType.choices,
                "revisions": self._revision_rows(emp),
            }
        )
        return context

    def _revision_rows(self, emp) -> list[dict]:
        rows = []
        qs = (
            SalaryRevision.objects.filter(user=emp)
            .select_related("approved_by")
            .order_by("-effective_date", "-created_at")[:10]
        )
        for rev in qs:
            pct = None
            if rev.previous_ctc:
                pct = round((rev.new_ctc - rev.previous_ctc) / rev.previous_ctc * 100, 1)
            rows.append({"rev": rev, "pct": pct})
        return rows


# ── New dedicated Payroll pages (sidebar restructure) ──────────────────────────


class PayrollDashboardView(PlanFeatureRequiredMixin, PayrollActionMixin, TemplateView):
    """Payroll → Dashboard: KPI cards, charts, quick actions, recent runs."""

    template_name = "payroll/payroll_dashboard.html"
    required_feature = "payroll_basic"
    action_redirect_url_name = "payroll:dashboard"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            org = request.user.organization
            if org:
                ensure_payroll_setup(org)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        org = request.user.organization
        result = self.post_payroll_action(request, org)
        if result is not None:
            return result
        messages.error(request, "Invalid action.")
        return self._redirect_back(request)

    def get_context_data(self, **kwargs):
        from .dashboard_analytics import dashboard_kpis, upcoming_payroll_date
        from .analytics import PayrollFilters, build_charts, recent_runs

        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = user.organization
        filters = PayrollFilters.from_request(self.request)

        kpis = dashboard_kpis(user, filters)
        runs = recent_runs(org, limit=5)
        context.update({
            "organization": org,
            "filters": filters,
            "kpis": kpis,
            "kpis_json": json.dumps(kpis),
            "charts_json": json.dumps(build_charts(user, filters)),
            "recent_runs": [
                {
                    "period_label": r.period_label,
                    "status": r.status,
                    "status_display": r.get_status_display(),
                    "employee_count": r.employee_count,
                    "net": abs(r.total_net),
                    "id": str(r.pk),
                }
                for r in runs
            ],
            "upcoming_payroll_date": upcoming_payroll_date(org),
            "is_finance": user.role in (User.Role.ADMIN, User.Role.HR),
            "is_admin": user.role == User.Role.ADMIN,
            "today": timezone.localdate(),
            "period_months": (_pm := _get_period_months(org)),
            "period_years": _get_period_years(_pm),
            "month_choices": [(i, calendar.month_name[i]) for i in range(1, 13)],
        })
        if user.role == User.Role.EMPLOYEE:
            # Employee dashboard embeds the My Salary structure below the KPI cards.
            context.update(_employee_salary_context(user, filters))
        return context


class PayrollCyclesView(PlanFeatureRequiredMixin, AdminOrHRRequiredMixin, TemplateView):
    """Payroll → Payroll Cycles: cycle frequency + cutoff configuration."""

    template_name = "payroll/cycles.html"
    required_feature = "payroll_basic"

    def post(self, request, *args, **kwargs):
        from .forms import PayrollCycleConfigForm
        from .models import PayrollCycleConfig
        from .services import record_payroll_action
        from .models import PayrollAuditLog

        org = request.user.organization
        cfg, _ = PayrollCycleConfig.objects.get_or_create(organization=org)
        form = PayrollCycleConfigForm(request.POST, instance=cfg)
        if form.is_valid():
            form.save()
            record_payroll_action(
                org, request.user, PayrollAuditLog.Action.SETTINGS_UPDATED,
                "Payroll cycle configuration updated", request=request,
            )
            messages.success(request, "Payroll cycle configuration saved.")
        else:
            messages.error(request, "Please correct the errors below.")
        return redirect("payroll:cycles")

    def get_context_data(self, **kwargs):
        from .forms import PayrollCycleConfigForm
        from .models import PayrollCycleConfig

        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        cfg, _ = PayrollCycleConfig.objects.get_or_create(organization=org)
        context.update({
            "organization": org,
            "cycle_config": cfg,
            "form": PayrollCycleConfigForm(instance=cfg),
        })
        return context


class PayrollRunsView(PlanFeatureRequiredMixin, AdminOrHRRequiredMixin, PayrollActionMixin, TemplateView):
    """Payroll → Payroll Runs: full run history + Draft→Calculate→Review→Approve→Paid→Lock workflow."""

    template_name = "payroll/runs.html"
    required_feature = "payroll_basic"
    action_redirect_url_name = "payroll:runs"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "bank":
            return self._export_bank_file(request)
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        org = request.user.organization
        result = self.post_payroll_action(request, org)
        if result is not None:
            return result
        messages.error(request, "Invalid action.")
        return self._redirect_back(request)

    def _export_bank_file(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        org = request.user.organization
        filters = self._filters_from_request(request)
        run = PayrollRun.objects.filter(
            organization=org, year=filters.year, month=filters.month
        ).first()
        if run is None:
            messages.error(
                request,
                "No payroll run exists for the selected period, so there is nothing to export.",
            )
            return redirect(f"{reverse('payroll:runs')}?year={filters.year}&month={filters.month}")

        org_name = getattr(org, "name", None) or "Organization"
        headers = [
            "S.No", "Employee ID", "Employee Name", "Bank Name",
            "Account Number", "IFSC Code", "Net Salary (INR)",
        ]
        total_cols = len(headers)
        last_col = get_column_letter(total_cols)

        # Palette
        NAVY = "1E3A8A"
        NAVY_DARK = "172554"
        HEAD = "2563EB"
        STRIPE = "EFF6FF"
        center = Alignment(horizontal="center", vertical="center")
        left_mid = Alignment(horizontal="left", vertical="center")
        right_mid = Alignment(horizontal="right", vertical="center")
        thin = Side(style="thin", color="BFDBFE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = '"₹" #,##0.00'

        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Transfer"

        # --- Banner ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title = ws.cell(row=1, column=1, value=org_name)
        title.font = Font(bold=True, size=18, color="FFFFFF")
        title.alignment = center
        title.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        sub = ws.cell(row=2, column=1, value="Salary Bank Transfer Advice")
        sub.font = Font(bold=True, size=12, color="FFFFFF")
        sub.alignment = center
        sub.fill = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

        # --- Meta line ---
        status = run.get_status_display() if hasattr(run, "get_status_display") else ""
        meta_bits = [
            f"Pay Period: {run.period_label}",
            f"Status: {status}" if status else None,
            f"Generated: {timezone.localdate():%d %b %Y}",
        ]
        meta_text = "        ".join(b for b in meta_bits if b)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
        meta = ws.cell(row=3, column=1, value=meta_text)
        meta.font = Font(bold=True, size=11, color=NAVY_DARK)
        meta.alignment = center
        meta.fill = PatternFill(start_color=STRIPE, end_color=STRIPE, fill_type="solid")
        meta.border = border
        ws.row_dimensions[3].height = 20

        # --- Table header ---
        header_row = 5
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color=HEAD, end_color=HEAD, fill_type="solid")
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[header_row].height = 22

        # --- Data rows ---
        current_row = header_row
        total_net = 0
        for idx, slip in enumerate(
            run.payslips.select_related("user").order_by("user__first_name", "user__last_name"),
            start=1,
        ):
            u = slip.user
            net = abs(slip.net_salary or 0)
            total_net += net
            current_row += 1
            values = [
                idx,
                u.employee_id or "",
                u.display_name,
                getattr(u, "bank_name", "") or "",
                str(getattr(u, "bank_account_number", "") or ""),
                getattr(u, "ifsc_code", "") or "",
                float(net),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                if col_idx == total_cols:
                    cell.alignment = right_mid
                    cell.number_format = money_fmt
                elif col_idx in (1,):
                    cell.alignment = center
                else:
                    cell.alignment = left_mid
            if idx % 2 == 0:  # zebra striping
                for col_idx in range(1, total_cols + 1):
                    ws.cell(row=current_row, column=col_idx).fill = PatternFill(
                        start_color=STRIPE, end_color=STRIPE, fill_type="solid"
                    )

        # --- Total row ---
        current_row += 1
        ws.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols - 1
        )
        tlabel = ws.cell(row=current_row, column=1, value="TOTAL")
        tlabel.font = Font(bold=True, color="FFFFFF", size=11)
        tlabel.alignment = right_mid
        tlabel.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        tval = ws.cell(row=current_row, column=total_cols, value=float(total_net))
        tval.font = Font(bold=True, color="FFFFFF", size=11)
        tval.alignment = right_mid
        tval.number_format = money_fmt
        tval.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        for col_idx in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col_idx).border = border

        # --- Footer note ---
        current_row += 2
        note = ws.cell(
            row=current_row,
            column=1,
            value=(
                f"Total {run.payslips.count()} employee(s) · This advice is system-generated "
                f"and valid without signature."
            ),
        )
        note.font = Font(size=8, italic=True, color="64748B")
        note.alignment = left_mid

        # --- Column widths + freeze ---
        widths = [6, 14, 26, 24, 22, 14, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{header_row + 1}"

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="bank_transfer_{run.year}{run.month:02d}.xlsx"'
        )
        wb.save(resp)
        return resp

    def get_context_data(self, **kwargs):
        from .analytics import PayrollFilters, approval_steps, get_current_run, recent_runs

        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        fy = get_selected_fy(self.request)
        filters = PayrollFilters.from_request(self.request, fy=fy)
        run = get_current_run(org, filters)
        runs = recent_runs(org, fy=fy)

        context.update({
            "organization": org,
            "filters": filters,
            "selected_fy": fy,
            "payroll_run": run,
            "approval_steps": approval_steps(run),
            "period_months": (_pm := _get_period_months(org, fy=fy)),
            "period_years": _get_period_years(_pm),
            "month_choices": [(i, calendar.month_name[i]) for i in range(1, 13)],
            "today": timezone.localdate(),
            "is_admin": self.request.user.role == User.Role.ADMIN,
            "run_history": [
                {
                    "year": r.year, "month": r.month, "period_label": r.period_label,
                    "employee_count": r.employee_count,
                    "gross": abs(r.total_gross), "deductions": abs(r.total_deductions),
                    "net": abs(r.total_net), "status": r.status,
                    "status_display": r.get_status_display(),
                }
                for r in runs
            ],
        })
        return context


class PayslipsView(PlanFeatureRequiredMixin, TemplateView):
    """Payroll → Payslips: browse/filter, single download, bulk generate/email/download."""

    template_name = "payroll/payslips_list.html"
    required_feature = "payroll_basic"
    paginate_by = 25

    def get(self, request, *args, **kwargs):
        if request.GET.get("payslip"):
            return self._payslip_pdf(request)
        return super().get(request, *args, **kwargs)

    def _payslip_pdf(self, request):
        from django.db.models import F

        from .models import Payslip
        from .pdf import render_payslip_pdf

        slip = get_object_or_404(
            Payslip.objects.select_related("user", "payroll_run").prefetch_related("lines"),
            pk=request.GET.get("payslip"),
            user__organization=request.user.organization,
        )
        if request.user.role not in (User.Role.ADMIN, User.Role.HR) and slip.user_id != request.user.pk:
            messages.error(request, "You can only view your own payslip.")
            return redirect("payroll:payslips")
        Payslip.objects.filter(pk=slip.pk).update(
            download_count=F("download_count") + 1, last_downloaded_at=timezone.now()
        )
        pdf_bytes = render_payslip_pdf(slip)
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        disposition = "inline" if request.GET.get("view") == "1" else "attachment"
        resp["Content-Disposition"] = f'{disposition}; filename="payslip_{slip.payslip_number or slip.pk.hex[:8]}.pdf"'
        return resp

    def post(self, request, *args, **kwargs):
        from .models import PayrollAuditLog
        from .services import generate_payslip_numbers, get_or_create_payroll_run, record_payroll_action

        org = request.user.organization
        action = request.POST.get("action")
        if action == "bulk_generate" and request.user.role in (User.Role.ADMIN, User.Role.HR):
            year = int(request.POST.get("year") or timezone.localdate().year)
            month = int(request.POST.get("month") or timezone.localdate().month)
            run = get_or_create_payroll_run(org, year, month)
            count = generate_payslip_numbers(run)
            record_payroll_action(
                org, request.user, PayrollAuditLog.Action.GENERATED,
                f"Bulk-generated {count} payslip(s) for {run.period_label}",
                period=run.period_label, request=request, count=count,
            )
            messages.success(request, f"Generated {count} payslip(s).")
        elif action == "bulk_email" and request.user.role in (User.Role.ADMIN, User.Role.HR):
            self._bulk_email(request, org)
        else:
            messages.error(request, "Invalid action.")
        return redirect("payroll:payslips")

    def _bulk_email(self, request, org):
        from apps.dashboard.notification_service import send_notification

        from .analytics import PayrollFilters, filtered_payslips_for_fy

        fy = get_selected_fy(request)
        filters = PayrollFilters.from_request(request, fy=fy)
        qs = filtered_payslips_for_fy(request.user, fy, filters)
        count = 0
        for slip in qs.select_related("user"):
            send_notification(
                slip.user,
                source_key=f"payslip_emailed_{slip.pk}",
                title="Payslip ready",
                message=f"Your payslip for {slip.payroll_run.period_label} is ready to download.",
                url=reverse("payroll:payslips"),
                icon="mail",
                force_unread=True,
            )
            count += 1
        messages.success(request, f"Notified {count} employee(s) about their payslip.")

    def get_context_data(self, **kwargs):
        from .analytics import PayrollFilters, filtered_payslips_for_fy, table_rows

        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = user.organization
        fy = get_selected_fy(self.request)
        filters = PayrollFilters.from_request(self.request, fy=fy)
        qs = filtered_payslips_for_fy(user, fy, filters)
        paginator = Paginator(table_rows(qs), self.paginate_by)
        page = paginator.get_page(self.request.GET.get("page") or 1)

        context.update({
            "organization": org,
            "filters": filters,
            "selected_fy": fy,
            "page_obj": page,
            "table_rows": page.object_list,
            "is_finance": user.role in (User.Role.ADMIN, User.Role.HR),
            "today": timezone.localdate(),
        })
        return context


class SalaryComponentsView(PlanFeatureRequiredMixin, AdminOrHRRequiredMixin, TemplateView):
    """Payroll → Salary Components: reusable earning/deduction component catalog."""

    template_name = "payroll/components.html"
    required_feature = "payroll_basic"

    def post(self, request, *args, **kwargs):
        from .forms import SalaryComponentForm
        from .models import PayrollAuditLog, SalaryComponent
        from .services import ensure_payroll_setup, record_payroll_action

        org = request.user.organization
        ensure_payroll_setup(org)
        action = request.POST.get("action")

        if action == "delete":
            comp = get_object_or_404(SalaryComponent, pk=request.POST.get("component_id"), organization=org)
            comp.is_active = False
            comp.save(update_fields=["is_active"])
            messages.success(request, f"Deactivated {comp.name}.")
            return redirect("payroll:components")

        instance = None
        comp_id = request.POST.get("component_id")
        if comp_id:
            instance = get_object_or_404(SalaryComponent, pk=comp_id, organization=org)
        form = SalaryComponentForm(request.POST, instance=instance)
        if form.is_valid():
            comp = form.save(commit=False)
            comp.organization = org
            comp.save()
            record_payroll_action(
                org, request.user, PayrollAuditLog.Action.SETTINGS_UPDATED,
                f"Salary component {'updated' if comp_id else 'created'}: {comp.name}",
                request=request,
            )
            messages.success(request, f"Saved {comp.name}.")
        else:
            messages.error(request, "Please correct the errors below.")
        return redirect("payroll:components")

    def get_context_data(self, **kwargs):
        from .forms import SalaryComponentForm
        from .models import SalaryComponent
        from .services import ensure_payroll_setup

        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        ensure_payroll_setup(org)
        components = SalaryComponent.objects.filter(organization=org).order_by(
            "-is_active", "component_type", "sort_order", "name"
        )
        context.update({
            "organization": org,
            "components": components,
            "form": SalaryComponentForm(),
        })
        return context


class TaxManagementView(PlanFeatureRequiredMixin, AdminOrHRRequiredMixin, TemplateView):
    """Payroll → Tax Management: tax regime, standard deduction, tax slabs."""

    template_name = "payroll/tax_management.html"
    required_feature = "payroll_advanced"

    def post(self, request, *args, **kwargs):
        from .forms import TaxConfigurationForm
        from .models import TaxConfiguration, TaxSlab
        from .services import ensure_payroll_setup

        org = request.user.organization
        ensure_payroll_setup(org)
        cfg = TaxConfiguration.objects.filter(organization=org, is_active=True).first()
        action = request.POST.get("action")

        if action == "add_slab":
            TaxSlab.objects.create(
                tax_config=cfg,
                min_income=Decimal(request.POST.get("min_income") or "0"),
                max_income=Decimal(request.POST["max_income"]) if request.POST.get("max_income") else None,
                rate_percent=Decimal(request.POST.get("rate_percent") or "0"),
            )
            messages.success(request, "Tax slab added.")
        elif action == "delete_slab":
            TaxSlab.objects.filter(pk=request.POST.get("slab_id"), tax_config=cfg).delete()
            messages.success(request, "Tax slab removed.")
        else:
            form = TaxConfigurationForm(request.POST, instance=cfg)
            if form.is_valid():
                form.save()
                messages.success(request, "Tax configuration saved.")
            else:
                messages.error(request, "Please correct the errors below.")
        return redirect("payroll:tax_management")

    def get_context_data(self, **kwargs):
        from .forms import TaxConfigurationForm
        from .models import TaxConfiguration
        from .services import ensure_payroll_setup

        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        ensure_payroll_setup(org)
        cfg = TaxConfiguration.objects.filter(organization=org, is_active=True).first()
        context.update({
            "organization": org,
            "tax_config": cfg,
            "slabs": cfg.slabs.all() if cfg else [],
            "form": TaxConfigurationForm(instance=cfg),
        })
        return context


class LoansAdvancesView(PlanFeatureRequiredMixin, TemplateView):
    """Payroll → Loans & Advances: employee loans with an approval workflow."""

    template_name = "payroll/loans.html"
    required_feature = "payroll_growth"

    def post(self, request, *args, **kwargs):
        from apps.dashboard.notification_service import send_notification

        from .forms import EmployeeLoanForm
        from .models import EmployeeLoan, PayrollAuditLog
        from .services import payroll_team_for, record_payroll_action

        user = request.user
        org = user.organization
        action = request.POST.get("action")

        if action == "apply" and user.role in (User.Role.HR, User.Role.EMPLOYEE):
            form = EmployeeLoanForm(request.POST)
            if form.is_valid():
                loan = form.save(commit=False)
                loan.user = user
                loan.balance = loan.principal
                loan.save()
                messages.success(request, "Loan application submitted for approval.")
            else:
                messages.error(request, "Please correct the errors below.")
        elif action == "approve" and user.role in (User.Role.ADMIN, User.Role.HR):
            loan = get_object_or_404(
                EmployeeLoan, pk=request.POST.get("loan_id"), user__in=payroll_team_for(user)
            )
            loan.status = EmployeeLoan.Status.ACTIVE
            loan.approved_by = user
            loan.approved_at = timezone.now()
            loan.save(update_fields=["status", "approved_by", "approved_at"])
            record_payroll_action(
                org, user, PayrollAuditLog.Action.LOAN_APPROVED,
                f"Loan approved for {loan.user.display_name}", request=request,
            )
            send_notification(
                loan.user, source_key=f"loan_approved_{loan.pk}", title="Loan approved",
                message=f"Your loan of {loan.principal} has been approved.",
                url=reverse("payroll:loans"), icon="hand-coins",
            )
            messages.success(request, "Loan approved.")
        elif action == "reject" and user.role in (User.Role.ADMIN, User.Role.HR):
            loan = get_object_or_404(
                EmployeeLoan, pk=request.POST.get("loan_id"), user__in=payroll_team_for(user)
            )
            loan.status = EmployeeLoan.Status.REJECTED
            loan.save(update_fields=["status"])
            messages.warning(request, "Loan rejected.")
        else:
            messages.error(request, "Invalid action.")
        return redirect("payroll:loans")

    def get_context_data(self, **kwargs):
        from .forms import EmployeeLoanForm
        from .models import EmployeeLoan
        from .services import payroll_team_for

        context = super().get_context_data(**kwargs)
        user = self.request.user
        team = payroll_team_for(user)
        loans = EmployeeLoan.objects.filter(user__in=team).select_related("user").order_by("-created_at")
        context.update({
            "organization": user.organization,
            "loans": loans,
            "form": EmployeeLoanForm(),
            "is_finance": user.role in (User.Role.ADMIN, User.Role.HR),
        })
        return context


class ReimbursementsView(PlanFeatureRequiredMixin, TemplateView):
    """Payroll → Reimbursements: claim / approve / reject expense reimbursements."""

    template_name = "payroll/reimbursements.html"
    required_feature = "payroll_growth"

    def post(self, request, *args, **kwargs):
        from .forms import ReimbursementForm
        from .models import Reimbursement
        from .services import payroll_team_for

        user = request.user
        action = request.POST.get("action")

        if action == "add":
            if user.role == User.Role.ADMIN:
                messages.error(request, "Admins approve reimbursements; submit as HR or employee.")
                return redirect("payroll:reimbursements")
            form = ReimbursementForm(request.POST, request.FILES)
            if form.is_valid():
                reimb = form.save(commit=False)
                reimb.user = user
                reimb.save()
                messages.success(request, "Reimbursement claim submitted.")
            else:
                messages.error(request, "Invalid reimbursement form.")
        elif action in ("approve", "reject") and user.role in (User.Role.ADMIN, User.Role.HR):
            reimb = get_object_or_404(
                Reimbursement, pk=request.POST.get("reimbursement_id"), user__in=payroll_team_for(user)
            )
            reimb.status = Reimbursement.Status.APPROVED if action == "approve" else Reimbursement.Status.REJECTED
            reimb.reviewed_by = user
            reimb.reviewed_at = timezone.now()
            reimb.save()
            messages.success(request, f"Reimbursement {action}d.")
        else:
            messages.error(request, "Invalid action.")
        return redirect("payroll:reimbursements")

    def get_context_data(self, **kwargs):
        from .analytics import pending_reimbursements

        from .forms import ReimbursementForm
        from .models import Reimbursement
        from .services import payroll_team_for

        context = super().get_context_data(**kwargs)
        user = self.request.user
        team = payroll_team_for(user)
        claims = Reimbursement.objects.filter(user__in=team).select_related("user").order_by("-created_at")
        context.update({
            "organization": user.organization,
            "claims": claims,
            "pending": pending_reimbursements(user),
            "form": ReimbursementForm(),
            "is_finance": user.role in (User.Role.ADMIN, User.Role.HR),
            "can_claim": user.role in (User.Role.HR, User.Role.EMPLOYEE),
        })
        return context


class SalaryRevisionsView(PlanFeatureRequiredMixin, AdminOrHRRequiredMixin, TemplateView):
    """Payroll → Salary Revisions: org-wide revision history."""

    template_name = "payroll/revisions.html"
    required_feature = "payroll_advanced"

    def get_context_data(self, **kwargs):
        from .models import SalaryRevision
        from .services import payroll_team_for

        context = super().get_context_data(**kwargs)
        user = self.request.user
        team = payroll_team_for(user)
        revisions = (
            SalaryRevision.objects.filter(user__in=team)
            .select_related("user", "approved_by")
            .order_by("-effective_date", "-created_at")
        )
        rows = []
        for rev in revisions:
            pct = None
            if rev.previous_ctc:
                pct = round((rev.new_ctc - rev.previous_ctc) / rev.previous_ctc * 100, 1)
            rows.append({"rev": rev, "pct": pct})
        context.update({"organization": user.organization, "revisions": rows})
        return context


class PayrollSettingsView(PlanFeatureRequiredMixin, AdminRequiredMixin, TemplateView):
    """Payroll → Payroll Settings: currency, rounding, auto-payroll, lock policy, defaults."""

    template_name = "payroll/settings.html"
    required_feature = "payroll_basic"

    def post(self, request, *args, **kwargs):
        from apps.organizations.models import Organization

        from .forms import PayrollSettingsForm
        from .models import PayrollAuditLog, PayrollSettings
        from .services import record_payroll_action

        org = request.user.organization
        settings_obj, _ = PayrollSettings.objects.get_or_create(organization=org)
        form = PayrollSettingsForm(request.POST, instance=settings_obj, organization=org)

        # The LOP policy lives on Organization, not on PayrollSettings, and is
        # validated on its own. Apply it independently of the form: it decides how
        # unmarked days are paid, and silently dropping it because an unrelated
        # PayrollSettings field failed validation loses a pay-affecting change.
        policy = request.POST.get("payroll_lop_policy")
        policy_saved = False
        if policy is not None:
            if policy in Organization.PayrollLopPolicy.values:
                if org.payroll_lop_policy != policy:
                    org.payroll_lop_policy = policy
                    org.save(update_fields=["payroll_lop_policy", "updated_at"])
                policy_saved = True
            else:
                messages.error(request, "Invalid payroll pay policy.")

        if form.is_valid():
            form.save()
            record_payroll_action(
                org, request.user, PayrollAuditLog.Action.SETTINGS_UPDATED,
                "Payroll settings updated", request=request,
            )
            messages.success(request, "Payroll settings saved.")
        else:
            if policy_saved:
                record_payroll_action(
                    org, request.user, PayrollAuditLog.Action.SETTINGS_UPDATED,
                    "Payroll pay policy updated", request=request,
                )
                messages.success(request, "Payroll pay policy saved.")
            messages.error(request, "Please correct the errors below.")
        return redirect("payroll:settings")

    def get_context_data(self, **kwargs):
        from apps.organizations.models import Organization

        from .forms import PayrollSettingsForm
        from .models import PayrollSettings

        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        settings_obj, _ = PayrollSettings.objects.get_or_create(organization=org)
        context.update({
            "organization": org,
            "settings": settings_obj,
            "form": PayrollSettingsForm(instance=settings_obj, organization=org),
            "lop_policy": getattr(org, "payroll_lop_policy", ""),
            "lop_policy_choices": Organization.PayrollLopPolicy.choices,
        })
        return context


def _salary_revision_rows(emp) -> list[dict]:
    rows = []
    qs = (
        SalaryRevision.objects.filter(user=emp)
        .order_by("-effective_date", "-created_at")[:10]
    )
    for rev in qs:
        pct = None
        if rev.previous_ctc:
            pct = round((rev.new_ctc - rev.previous_ctc) / rev.previous_ctc * 100, 1)
        rows.append({"rev": rev, "pct": pct})
    return rows


def _employee_salary_context(emp, filters) -> dict:
    """Employee's own salary breakdown for the selected period (payslip if one
    exists, otherwise the active salary structure) + revision history."""
    from .services import compute_employee_breakdown, get_active_salary, payslip_breakdown

    salary = get_active_salary(emp)
    payslip = (
        Payslip.objects.filter(
            user=emp,
            payroll_run__organization=emp.organization,
            payroll_run__year=filters.year,
            payroll_run__month=filters.month,
        )
        .select_related("payroll_run")
        .prefetch_related("lines")
        .first()
    )
    breakdown = payslip_breakdown(payslip) if payslip else compute_employee_breakdown(salary)
    return {
        "salary": salary,
        "payslip": payslip,
        "breakdown": breakdown,
        "revisions": _salary_revision_rows(emp),
    }


class PayrollPlaceholderView(PlanFeatureRequiredMixin, TemplateView):
    """Generic 'coming in a future release' page for Form 16 / Bonuses / Overtime /
    Arrears / Final Settlement — real implementations are separate follow-up phases."""

    template_name = "payroll/coming_soon.html"
    required_feature = "payroll_growth"
    feature_name = "This feature"
    feature_description = "This module is coming in a future release."

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            "organization": self.request.user.organization,
            "feature_name": self.feature_name,
            "feature_description": self.feature_description,
        })
        return context
