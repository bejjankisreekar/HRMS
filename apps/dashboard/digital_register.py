"""Digital Attendance Register — register-book style monthly grid builder.

Produces a per-employee × per-day status grid for ADMIN/HR, reusing the
existing attendance, weekend-policy and holiday infrastructure.
"""
from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date, timedelta

from django.utils import timezone

from apps.accounts.hierarchy import attendance_team_for
from apps.accounts.models import User
from apps.attendance.models import AttendanceRecord
from apps.attendance.work_calendar import (
    get_holidays_between,
    get_org_off_weekdays,
    get_shift_off_weekdays,
)
from apps.organizations.models import Department, Organization

from .attendance_utils import analyze_lateness, get_effective_shift

# Register status codes
CODE_PRESENT = "P"
CODE_ABSENT = "A"
CODE_LEAVE = "L"
CODE_HALF = "HD"
CODE_WFH = "WFH"
CODE_HOLIDAY = "H"
CODE_WEEKEND = "WO"
CODE_LATE = "LT"
CODE_BLANK = ""  # not marked, working day

STATUS_META = {
    CODE_PRESENT: {"label": "Present", "cls": "dr-present"},
    CODE_ABSENT: {"label": "Absent", "cls": "dr-absent"},
    CODE_LEAVE: {"label": "Leave", "cls": "dr-leave"},
    CODE_HALF: {"label": "Half day", "cls": "dr-half"},
    CODE_WFH: {"label": "Work from home", "cls": "dr-wfh"},
    CODE_HOLIDAY: {"label": "Holiday", "cls": "dr-holiday"},
    CODE_WEEKEND: {"label": "Weekly off", "cls": "dr-weekend"},
    CODE_LATE: {"label": "Late", "cls": "dr-late"},
    CODE_BLANK: {"label": "Not marked", "cls": "dr-blank"},
}

LEGEND = [
    (CODE_PRESENT, "Present"),
    (CODE_LATE, "Late"),
    (CODE_HALF, "Half day"),
    (CODE_WFH, "WFH"),
    (CODE_LEAVE, "Leave"),
    (CODE_ABSENT, "Absent"),
    (CODE_HOLIDAY, "Holiday"),
    (CODE_WEEKEND, "Weekly off"),
]


@dataclass
class RegisterFilters:
    start: date
    end: date
    department_id: str = ""
    employee_id: str = ""
    search: str = ""

    @classmethod
    def from_request(cls, request) -> "RegisterFilters":
        today = timezone.localdate()
        g = request.GET

        # Custom range takes priority if both provided
        start_raw = g.get("start_date") or ""
        end_raw = g.get("end_date") or ""
        if start_raw and end_raw:
            start = _parse_date(start_raw) or today.replace(day=1)
            end = _parse_date(end_raw) or _month_end(start)
        else:
            try:
                month = int(g.get("month") or today.month)
                year = int(g.get("year") or today.year)
            except (TypeError, ValueError):
                month, year = today.month, today.year
            month = min(max(month, 1), 12)
            start = date(year, month, 1)
            end = _month_end(start)

        # Cap range to a sane maximum (62 days) to protect rendering
        if (end - start).days > 62:
            end = start + timedelta(days=62)

        return cls(
            start=start,
            end=end,
            department_id=g.get("department") or g.get("department_id") or "",
            employee_id=g.get("employee") or g.get("employee_id") or "",
            search=(g.get("search") or "").strip(),
        )


def _parse_date(raw: str) -> date | None:
    try:
        y, m, d = (int(x) for x in raw.split("-"))
        return date(y, m, d)
    except (ValueError, AttributeError):
        return None


def _month_end(d: date) -> date:
    last = calendar.monthrange(d.year, d.month)[1]
    return date(d.year, d.month, last)


def _daterange(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def _resolve_team(manager: User, filters: RegisterFilters):
    qs = attendance_team_for(manager).select_related("work_shift", "department")
    if filters.department_id:
        qs = qs.filter(department_id=filters.department_id)
    if filters.employee_id:
        qs = qs.filter(pk=filters.employee_id)
    if filters.search:
        from django.db.models import Q

        term = filters.search
        qs = qs.filter(
            Q(first_name__icontains=term)
            | Q(last_name__icontains=term)
            | Q(username__icontains=term)
            | Q(employee_id__icontains=term)
        )
    return qs.order_by("first_name", "last_name", "username")


def build_register(manager: User, filters: RegisterFilters, *, page: int = 1, page_size: int = 100) -> dict:
    org: Organization = manager.organization
    start, end = filters.start, filters.end
    today = timezone.localdate()

    days = list(_daterange(start, end))
    day_headers = [
        {
            "date": d,
            "day": d.day,
            "weekday": d.strftime("%a"),
            "is_today": d == today,
            "is_weekend": d.weekday() >= 5,
            "is_week_start": d.weekday() == 0,
            "iso": d.isoformat(),
        }
        for d in days
    ]

    team_qs = _resolve_team(manager, filters)
    total_employees = team_qs.count()

    # Server-side pagination
    page = max(1, page)
    offset = (page - 1) * page_size
    team = list(team_qs[offset : offset + page_size])
    team_ids = [u.pk for u in team]

    # Holidays in range (set of dates)
    holidays = get_holidays_between(org, start, end) if org else set()

    # Weekend weekdays per date (org-level). Shift-based handled per-user below.
    org_off_by_date: dict[date, set[int]] = {}
    if org:
        for d in days:
            org_off_by_date[d] = get_org_off_weekdays(org, d)
    shift_based = bool(
        org and org.weekend_policy == Organization.WeekendPolicy.SHIFT_BASED
    )

    # Bulk-load attendance records for the team in range
    records = AttendanceRecord.objects.filter(
        user_id__in=team_ids, date__gte=start, date__lte=end
    )
    rec_map: dict[tuple, AttendanceRecord] = {}
    for rec in records:
        rec_map[(rec.user_id, rec.date)] = rec

    rows = []
    for member in team:
        shift = get_effective_shift(member)
        # per-user weekend resolution (cached by shift for shift-based policy)
        user_off_cache: dict[date, set[int]] = {}
        if shift_based:
            for d in days:
                off = get_shift_off_weekdays(member, d)
                user_off_cache[d] = off or {5, 6}

        cells = []
        present = absent = leave = half = wfh = late = holiday_ct = weekend_ct = 0

        for d in days:
            rec = rec_map.get((member.pk, d))
            is_weekend = (
                d.weekday() in user_off_cache.get(d, set())
                if shift_based
                else d.weekday() in org_off_by_date.get(d, set())
            )
            is_holiday = d in holidays

            code = CODE_BLANK
            tip = None

            if rec and rec.status == AttendanceRecord.Status.PRESENT:
                lateness = analyze_lateness(rec, shift, d)
                code = CODE_LATE if lateness.get("is_late") else CODE_PRESENT
                present += 1
                if code == CODE_LATE:
                    late += 1
            elif rec and rec.status == AttendanceRecord.Status.HALF_DAY:
                code = CODE_HALF
                half += 1
            elif rec and rec.status == AttendanceRecord.Status.WFH:
                code = CODE_WFH
                wfh += 1
            elif rec and rec.status == AttendanceRecord.Status.LEAVE:
                code = CODE_LEAVE
                leave += 1
            elif rec and rec.status == AttendanceRecord.Status.ABSENT:
                code = CODE_ABSENT
                absent += 1
            elif is_holiday:
                code = CODE_HOLIDAY
                holiday_ct += 1
            elif is_weekend:
                code = CODE_WEEKEND
                weekend_ct += 1
            else:
                # no record on a working day
                code = CODE_BLANK

            if rec:
                tip = {
                    "in": _fmt_time(rec.check_in),
                    "out": _fmt_time(rec.check_out),
                    "hours": _fmt_hours(rec),
                }

            cells.append(
                {
                    "code": code,
                    "cls": STATUS_META[code]["cls"],
                    "label": STATUS_META[code]["label"],
                    "iso": d.isoformat(),
                    "day": d.day,
                    "is_weekend": is_weekend,
                    "is_week_start": d.weekday() == 0,
                    "is_today": d == today,
                    "tip": tip,
                }
            )

        rows.append(
            {
                "member": member,
                "employee_id": member.employee_id or "—",
                "name": member.display_name,
                "department": member.department.name if member.department_id else "—",
                "cells": cells,
                "present_days": present,
                "absent_days": absent,
                "leave_days": leave,
                "half_days": half,
                "wfh_days": wfh,
                "late_days": late,
            }
        )

    total_pages = max(1, (total_employees + page_size - 1) // page_size)

    return {
        "day_headers": day_headers,
        "rows": rows,
        "legend": LEGEND,
        "total_employees": total_employees,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "showing_from": offset + 1 if team else 0,
        "showing_to": offset + len(team),
        "start": start,
        "end": end,
        "num_days": len(days),
    }


def build_summary_cards(manager: User, filters: RegisterFilters) -> dict:
    org = manager.organization
    today = timezone.localdate()
    team_qs = _resolve_team(manager, filters)
    total = team_qs.count()
    team_ids = list(team_qs.values_list("pk", flat=True))

    todays = AttendanceRecord.objects.filter(user_id__in=team_ids, date=today)
    present = todays.filter(
        status__in=[AttendanceRecord.Status.PRESENT, AttendanceRecord.Status.WFH]
    ).count()
    absent = todays.filter(status=AttendanceRecord.Status.ABSENT).count()
    on_leave = todays.filter(status=AttendanceRecord.Status.LEAVE).count()

    # Attendance % over the selected range (present+wfh+half / marked working slots)
    range_recs = AttendanceRecord.objects.filter(
        user_id__in=team_ids, date__gte=filters.start, date__lte=filters.end
    )
    marked = range_recs.count()
    good = range_recs.filter(
        status__in=[
            AttendanceRecord.Status.PRESENT,
            AttendanceRecord.Status.WFH,
            AttendanceRecord.Status.HALF_DAY,
        ]
    ).count()
    pct = round(good / marked * 100, 1) if marked else 0.0

    return {
        "total_employees": total,
        "present_today": present,
        "absent_today": absent,
        "on_leave_today": on_leave,
        "attendance_pct": pct,
    }


def register_filter_options(manager: User) -> dict:
    org = manager.organization
    team = attendance_team_for(manager)
    departments = (
        Department.objects.filter(organization=org, is_active=True).order_by(
            "sort_order", "name"
        )
        if org
        else Department.objects.none()
    )
    return {
        "departments": list(departments),
        "employees": list(team.order_by("first_name", "last_name")),
        "dept_label": org.department_label if org else "Department",
        "dept_label_plural": org.department_label_plural if org else "Departments",
    }


def _fmt_time(dt) -> str:
    if not dt:
        return "—"
    return timezone.localtime(dt).strftime("%I:%M %p").lstrip("0")


def _fmt_hours(rec: AttendanceRecord) -> str:
    if not rec.check_in or not rec.check_out:
        return "—"
    mins = int((rec.check_out - rec.check_in).total_seconds() // 60) - (rec.break_minutes or 0)
    if mins <= 0:
        return "—"
    h, m = divmod(mins, 60)
    return f"{h:02d}h {m:02d}m"


def build_register_json(manager: User, filters: RegisterFilters, page: int = 1) -> dict:
    """API-shaped payload matching the spec."""
    data = build_register(manager, filters, page=page)
    employees = []
    for row in data["rows"]:
        attendance = {}
        for cell, hdr in zip(row["cells"], data["day_headers"]):
            attendance[str(hdr["day"])] = cell["code"] or "-"
        employees.append(
            {
                "employee_id": row["employee_id"],
                "employee_name": row["name"],
                "department": row["department"],
                "attendance": attendance,
                "present_days": row["present_days"],
                "absent_days": row["absent_days"],
                "leave_days": row["leave_days"],
                "half_days": row["half_days"],
                "wfh_days": row["wfh_days"],
                "late_days": row["late_days"],
            }
        )
    return {
        "start_date": filters.start.isoformat(),
        "end_date": filters.end.isoformat(),
        "days": [h["day"] for h in data["day_headers"]],
        "employees": employees,
        "pagination": {
            "page": data["page"],
            "page_size": data["page_size"],
            "total_pages": data["total_pages"],
            "total_employees": data["total_employees"],
        },
    }


def export_register_csv(manager: User, filters: RegisterFilters):
    import csv
    from django.http import HttpResponse

    data = build_register(manager, filters, page=1, page_size=100000)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="attendance_register_{filters.start}_{filters.end}.csv"'
    )
    writer = csv.writer(response)
    header = ["Employee ID", "Employee Name", "Department"]
    header += [f"{h['day']} ({h['weekday']})" for h in data["day_headers"]]
    header += ["Present", "Absent", "Leave", "Half", "WFH", "Late"]
    writer.writerow(header)
    for row in data["rows"]:
        line = [row["employee_id"], row["name"], row["department"]]
        line += [c["code"] or "-" for c in row["cells"]]
        line += [
            row["present_days"],
            row["absent_days"],
            row["leave_days"],
            row["half_days"],
            row["wfh_days"],
            row["late_days"],
        ]
        writer.writerow(line)
    return response


def export_register_xlsx(manager: User, filters: RegisterFilters):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    data = build_register(manager, filters, page=1, page_size=100000)

    org = getattr(manager, "organization", None)
    org_name = org.name if org else "Organization"

    dept_label = ""
    if filters.department_id:
        dept = Department.objects.filter(pk=filters.department_id).first()
        if dept:
            dept_label = dept.name

    day_headers = data["day_headers"]
    num_days = len(day_headers)
    weekend_flags = [h["date"].weekday() == 6 for h in day_headers]  # Sunday only
    month_name = filters.start.strftime("%B")
    year_label = filters.start.strftime("%Y")
    summary_labels = ["Present", "Absent", "Leave", "Half", "WFH", "Late"]

    # Column layout: Emp ID | Name | Month | <days...> | <summary...>
    first_day_col = 4
    first_summary_col = first_day_col + num_days
    total_cols = first_summary_col + len(summary_labels) - 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # Palette (approximating the reference template)
    TEAL = "13A89C"
    TEAL_DARK = "0E8A80"
    WEEKEND = "FFF3CD"
    HEAD_LIGHT = "F1F5F9"
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center")
    last_col = get_column_letter(total_cols)

    teal_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
    teal_dark_fill = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
    weekend_fill = PatternFill(start_color=WEEKEND, end_color=WEEKEND, fill_type="solid")
    light_fill = PatternFill(start_color=HEAD_LIGHT, end_color=HEAD_LIGHT, fill_type="solid")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------------------------------------------------------------
    # Banner: company name + "Employee Attendance Sheet" (full width)
    # ---------------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title = ws.cell(row=1, column=1, value=org_name)
    title.font = Font(bold=True, size=22, color="FFFFFF")
    title.alignment = center
    title.fill = teal_fill

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    subtitle = ws.cell(row=2, column=1, value="Employee Attendance Sheet")
    subtitle.font = Font(bold=True, size=13, color="FFFFFF")
    subtitle.alignment = center
    subtitle.fill = teal_fill

    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 24

    # ---------------------------------------------------------------
    # Meta line: Year / Month — highlighted as the key context
    # ---------------------------------------------------------------
    meta_row = 3
    ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=total_cols)
    meta_bits = [f"Year: {year_label}", f"Month: {month_name}"]
    if dept_label:
        meta_bits.append(f"Department: {dept_label}")
    meta = ws.cell(row=meta_row, column=1, value="       ".join(meta_bits))
    meta.font = Font(bold=True, size=13, color=TEAL_DARK)
    meta.alignment = center
    meta.fill = PatternFill(start_color="D1F5F0", end_color="D1F5F0", fill_type="solid")
    meta.border = border
    ws.row_dimensions[meta_row].height = 24

    # ---------------------------------------------------------------
    # Two-row table header:
    #   row H1: identity labels (merged down) + day numbers + summary (merged down)
    #   row H2: weekday abbreviations under each day number
    # ---------------------------------------------------------------
    h1 = 4
    h2 = 5

    # Identity headers span both header rows
    for col_idx, label in ((1, "Emp. ID"), (2, "Employee Name"), (3, "Department")):
        ws.merge_cells(start_row=h1, start_column=col_idx, end_row=h2, end_column=col_idx)
        cell = ws.cell(row=h1, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = teal_dark_fill
        cell.alignment = center_wrap
        cell.border = border
        ws.cell(row=h2, column=col_idx).border = border

    # Day columns: number on top, weekday below
    for i, hdr in enumerate(day_headers):
        col_idx = first_day_col + i
        num_cell = ws.cell(row=h1, column=col_idx, value=hdr["day"])
        wk_cell = ws.cell(row=h2, column=col_idx, value=hdr["weekday"])
        fill = weekend_fill if weekend_flags[i] else light_fill
        for cell in (num_cell, wk_cell):
            cell.fill = fill
            cell.alignment = center
            cell.border = border
        num_cell.font = Font(bold=True, size=9, color="0F172A")
        wk_cell.font = Font(size=8, color="475569")

    # Summary headers span both header rows, vertical text
    vertical = Alignment(text_rotation=90, horizontal="center", vertical="center")
    for i, label in enumerate(summary_labels):
        col_idx = first_summary_col + i
        ws.merge_cells(start_row=h1, start_column=col_idx, end_row=h2, end_column=col_idx)
        cell = ws.cell(row=h1, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = teal_dark_fill
        cell.alignment = vertical
        cell.border = border
        ws.cell(row=h2, column=col_idx).border = border

    ws.row_dimensions[h1].height = 18
    ws.row_dimensions[h2].height = 44

    # ---------------------------------------------------------------
    # Data rows
    # ---------------------------------------------------------------
    current_row = h2
    for row in data["rows"]:
        current_row += 1
        line = [row["employee_id"], row["name"], row["department"]]
        line += [c["code"] or "-" for c in row["cells"]]
        line += [
            row["present_days"],
            row["absent_days"],
            row["leave_days"],
            row["half_days"],
            row["wfh_days"],
            row["late_days"],
        ]
        for col_idx, value in enumerate(line, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            if col_idx == 2:
                cell.alignment = left_mid
            else:
                cell.alignment = center
            # Shade weekend day-columns lightly
            if first_day_col <= col_idx < first_summary_col and weekend_flags[col_idx - first_day_col]:
                cell.fill = weekend_fill

    last_data_row = current_row

    # ---------------------------------------------------------------
    # Column widths + freeze panes
    # ---------------------------------------------------------------
    ws.freeze_panes = f"D{h1}"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    for col_idx in range(first_day_col, first_summary_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 3.6
    for col_idx in range(first_summary_col, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.5

    # ---------------------------------------------------------------
    # Legend (full forms of the status codes)
    # ---------------------------------------------------------------
    current_row += 2  # gap below the table
    legend_cell = ws.cell(row=current_row, column=1, value="Legend")
    legend_cell.font = Font(bold=True, size=11, color=TEAL_DARK)
    legend_items = list(LEGEND) + [("-", "Not marked")]
    for code, label in legend_items:
        current_row += 1
        code_cell = ws.cell(row=current_row, column=1, value=code)
        code_cell.font = Font(bold=True, color=TEAL_DARK)
        code_cell.alignment = center
        ws.cell(row=current_row, column=2, value=label).alignment = left_mid

    # Generated-on note (small, unobtrusive)
    current_row += 2
    gen = ws.cell(
        row=current_row,
        column=1,
        value=f"Generated on {timezone.localdate():%d %b %Y}",
    )
    gen.font = Font(size=8, italic=True, color="94A3B8")
    gen.alignment = left_mid

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="attendance_register_{filters.start}_{filters.end}.xlsx"'
    )
    wb.save(response)
    return response
