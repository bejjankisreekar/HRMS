"""HR Analytics — enterprise workforce intelligence layer.

Pure reporting module: it owns no models. Every metric is derived from the
existing accounts / attendance / leaves / payroll / lifecycle tables and is
returned as a JSON-serialisable dict for the HR Analytics dashboard.

Metric definitions follow standard HR practice:

  Attrition rate        separations / average headcount over the period
  Annualised attrition  period rate * (12 / months in period)
  Retention rate        100 - attrition rate
  Absenteeism rate      unplanned absent days / scheduled working days
  Average tenure        mean of (as-of date - joining date) for the population
  Span of control       active employees / people with at least one direct report
  Cost per employee     total payroll cost / paid headcount
  Gender pay gap        (male mean CTC - female mean CTC) / male mean CTC
"""

from __future__ import annotations

import calendar
import csv
import io
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any, Iterable

from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.utils import timezone

from apps.accounts.models import User
from apps.organizations.models import Department

WORKFORCE_ROLES = (User.Role.HR, User.Role.EMPLOYEE)

# Reasons that count as the employee choosing to leave.
VOLUNTARY_REASONS = {"BETTER_OPPORTUNITY", "PERSONAL", "RELOCATION", "HEALTH", "OTHER"}
INVOLUNTARY_REASONS = {"TERMINATION"}
# Retirement is a separation but neither regretted attrition nor a dismissal.
NEUTRAL_REASONS = {"RETIREMENT"}

TENURE_BANDS = (
    ("<6 months", 0.0, 0.5),
    ("6-12 months", 0.5, 1.0),
    ("1-2 years", 1.0, 2.0),
    ("2-5 years", 2.0, 5.0),
    ("5+ years", 5.0, 1000.0),
)

AGE_BANDS = (
    ("Under 25", 0.0, 25.0),
    ("25-34", 25.0, 35.0),
    ("35-44", 35.0, 45.0),
    ("45-54", 45.0, 55.0),
    ("55+", 55.0, 200.0),
)

PERIOD_CHOICES = (
    ("this_month", "This month"),
    ("last_month", "Last month"),
    ("this_quarter", "This quarter"),
    ("last_quarter", "Last quarter"),
    ("last_6m", "Last 6 months"),
    ("last_12m", "Last 12 months"),
    ("ytd", "Year to date"),
    ("fy", "Financial year"),
    ("custom", "Custom range"),
)

SECTIONS = (
    "overview", "workforce", "attrition", "attendance",
    "compensation", "diversity", "scorecard",
)

CACHE_TTL = 120


CURRENCY_SYMBOLS = {"INR": "₹", "USD": "$", "EUR": "€", "GBP": "£"}


def currency_symbol(org) -> str:
    code = (getattr(org, "currency", "") or "INR").upper()
    return CURRENCY_SYMBOLS.get(code, code + " ")


# -- small numeric helpers ----------------------------------------------------

def _pct(numerator: float, denominator: float, digits: int = 1) -> float:
    """Percentage, safe against a zero denominator."""
    if not denominator:
        return 0.0
    return round(numerator / denominator * 100, digits)


def _div(numerator: float, denominator: float, digits: int = 1) -> float:
    if not denominator:
        return 0.0
    return round(numerator / denominator, digits)


def _f(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _delta(current: float, previous: float | None, digits: int = 1) -> float | None:
    """Percentage change vs the comparison window. None when there is no base."""
    if not previous:
        return None
    return round((current - previous) / abs(previous) * 100, digits)


def _month_label(year: int, month: int) -> str:
    return f"{calendar.month_abbr[month]} {str(year)[2:]}"


def _month_end(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def _add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    total = (year * 12 + (month - 1)) + delta
    return total // 12, total % 12 + 1


def _add_months_date(d: date, delta: int) -> date:
    y, m = _add_months(d.year, d.month, delta)
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))


def _month_span(start: date, end: date) -> list[tuple[int, int]]:
    """Inclusive list of (year, month) buckets between two dates."""
    out: list[tuple[int, int]] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month) and len(out) < 120:
        out.append((y, m))
        y, m = _add_months(y, m, 1)
    return out


def _years_between(start: date | None, end: date) -> float | None:
    if not start:
        return None
    return max(0.0, (end - start).days / 365.25)


def _band_for(value: float | None, bands: Iterable[tuple[str, float, float]]) -> str | None:
    if value is None:
        return None
    for label, low, high in bands:
        if low <= value < high:
            return label
    return None


def _parse_date(raw: str | None) -> date | None:
    if not raw:
        return None
    try:
        return datetime.strptime(raw.strip(), "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        return None


def _top_n(counter: dict, limit: int = 12) -> list[tuple[Any, int]]:
    return sorted(counter.items(), key=lambda kv: kv[1], reverse=True)[:limit]


# -- filters ------------------------------------------------------------------

def resolve_period(period: str, today: date, fy: dict | None = None) -> tuple[date, date] | None:
    """Resolve a named period to (date_from, date_to). None = caller supplies dates."""
    if period == "this_month":
        return today.replace(day=1), today
    if period == "last_month":
        prev_end = today.replace(day=1) - timedelta(days=1)
        return prev_end.replace(day=1), prev_end
    if period == "this_quarter":
        q_month = 3 * ((today.month - 1) // 3) + 1
        return date(today.year, q_month, 1), today
    if period == "last_quarter":
        q_month = 3 * ((today.month - 1) // 3) + 1
        prev_end = date(today.year, q_month, 1) - timedelta(days=1)
        prev_q_month = 3 * ((prev_end.month - 1) // 3) + 1
        return date(prev_end.year, prev_q_month, 1), prev_end
    if period == "last_6m":
        y, m = _add_months(today.year, today.month, -5)
        return date(y, m, 1), today
    if period == "last_12m":
        y, m = _add_months(today.year, today.month, -11)
        return date(y, m, 1), today
    if period == "ytd":
        return date(today.year, 1, 1), today
    if period == "fy" and fy:
        start, end = fy.get("date_from"), fy.get("date_to")
        if isinstance(start, str):
            start = _parse_date(start)
        if isinstance(end, str):
            end = _parse_date(end)
        if start and end:
            return start, min(end, today)
    return None


@dataclass
class HRFilters:
    """Everything the dashboard can slice by."""

    date_from: date
    date_to: date
    period: str = "last_12m"
    department: str = ""
    employment_type: str = ""
    work_mode: str = ""
    location: str = ""
    fy_label: str = ""

    @classmethod
    def from_request(cls, request, fy: dict | None = None) -> "HRFilters":
        today = timezone.localdate()
        period = (request.GET.get("period") or "last_12m").strip()
        resolved = resolve_period(period, today, fy)
        if resolved:
            date_from, date_to = resolved
        else:
            period = "custom"
            date_from = _parse_date(request.GET.get("from")) or _add_months_date(today, -11).replace(day=1)
            date_to = _parse_date(request.GET.get("to")) or today
        if date_to < date_from:
            date_from, date_to = date_to, date_from
        return cls(
            date_from=date_from,
            date_to=date_to,
            period=period,
            department=(request.GET.get("department") or "").strip(),
            employment_type=(request.GET.get("employment_type") or "").strip(),
            work_mode=(request.GET.get("work_mode") or "").strip(),
            location=(request.GET.get("location") or "").strip(),
            fy_label=(fy or {}).get("label", "") if fy else "",
        )

    @property
    def months(self) -> list[tuple[int, int]]:
        return _month_span(self.date_from, self.date_to)

    @property
    def month_count(self) -> int:
        return max(1, len(self.months))

    @property
    def trend_months(self) -> list[tuple[int, int]]:
        """Month buckets for trend charts - never fewer than 12, for context."""
        span = self.months
        if len(span) >= 6:
            return span[-24:]
        y, m = _add_months(self.date_to.year, self.date_to.month, -11)
        return _month_span(date(y, m, 1), self.date_to)

    @property
    def previous_range(self) -> tuple[date, date]:
        """Equal-length window immediately before this one, for delta chips."""
        length = (self.date_to - self.date_from).days + 1
        prev_end = self.date_from - timedelta(days=1)
        return prev_end - timedelta(days=length - 1), prev_end

    def label(self) -> str:
        if self.period == "fy" and self.fy_label:
            return f"FY {self.fy_label}"
        mapping = dict(PERIOD_CHOICES)
        if self.period in mapping and self.period != "custom":
            return mapping[self.period]
        return f"{self.date_from:%d %b %Y} - {self.date_to:%d %b %Y}"

    def cache_suffix(self) -> str:
        return "|".join([
            self.date_from.isoformat(), self.date_to.isoformat(), self.period,
            self.department, self.employment_type, self.work_mode, self.location,
        ])

    def as_dict(self) -> dict:
        return {
            "period": self.period,
            "label": self.label(),
            "date_from": self.date_from.isoformat(),
            "date_to": self.date_to.isoformat(),
            "months": self.month_count,
            "department": self.department,
            "employment_type": self.employment_type,
            "work_mode": self.work_mode,
            "location": self.location,
        }


# -- workforce snapshot -------------------------------------------------------

@dataclass
class EmployeeRow:
    """One person, flattened - every metric in this module reads from these."""

    id: str
    name: str
    employee_code: str
    department_id: Any
    department: str
    designation: str
    grade: str
    grade_level: int | None
    gender: str
    employment_type: str
    work_mode: str
    employment_status: str
    location: str
    is_active: bool
    join_date: date | None
    exit_date: date | None
    exit_reason: str
    birth_date: date | None
    monthly_ctc: float
    manager_id: Any

    def active_on(self, on_date: date) -> bool:
        """Was this person on the payroll on the given date?"""
        if self.join_date and self.join_date > on_date:
            return False
        if self.exit_date and self.exit_date < on_date:
            return False
        if not self.join_date and not self.is_active:
            return False
        return True

    def tenure_years(self, as_of: date) -> float | None:
        end = min(self.exit_date, as_of) if self.exit_date else as_of
        return _years_between(self.join_date, end)

    def age(self, as_of: date) -> float | None:
        return _years_between(self.birth_date, as_of)


@dataclass
class Workforce:
    """Loaded-once population for a request, plus the lookup maps metrics need."""

    org: Any
    filters: HRFilters
    rows: list[EmployeeRow]
    departments: dict[Any, str]
    manager_report_counts: dict[Any, int]
    # Per-request memo so build_overview can compose the other sections cheaply.
    cache: dict = field(default_factory=dict)

    @property
    def today(self) -> date:
        return self.filters.date_to

    def active(self, on_date: date | None = None) -> list[EmployeeRow]:
        on_date = on_date or self.today
        return [r for r in self.rows if r.active_on(on_date)]

    def headcount(self, on_date: date | None = None) -> int:
        return len(self.active(on_date))

    def joiners(self, start: date, end: date) -> list[EmployeeRow]:
        return [r for r in self.rows if r.join_date and start <= r.join_date <= end]

    def leavers(self, start: date, end: date) -> list[EmployeeRow]:
        return [r for r in self.rows if r.exit_date and start <= r.exit_date <= end]

    def average_headcount(self, start: date, end: date) -> float:
        """Mean of the opening and closing headcount - the standard denominator."""
        opening = self.headcount(start)
        closing = self.headcount(end)
        return (opening + closing) / 2 or float(closing)

    def attrition_rate(self, start: date, end: date, annualised: bool = False) -> float:
        avg = self.average_headcount(start, end)
        if not avg:
            return 0.0
        rate = len(self.leavers(start, end)) / avg * 100
        if annualised:
            months = max(1, len(_month_span(start, end)))
            rate = rate * 12 / months
        return round(rate, 1)


def scheduled_capacity(org, members, start: date, end: date, wd_cache: dict | None = None) -> float:
    """Working days x headcount, summed month by month.

    One average headcount across a long window distorts the denominator for any
    org whose headcount moved, so each month is measured on its own. Pass
    ``wd_cache`` (a plain dict) to reuse the working-day counts across calls -
    each one is a holiday-table query, and the scorecard makes hundreds.
    """
    from apps.attendance.work_calendar import count_working_days

    cache_map = wd_cache if wd_cache is not None else {}
    total = 0.0
    for y, m in _month_span(start, end):
        m_start = max(date(y, m, 1), start)
        m_end = min(_month_end(y, m), end)
        if m_end < m_start:
            continue
        opening = sum(1 for r in members if r.active_on(m_start))
        closing = sum(1 for r in members if r.active_on(m_end))
        headcount = (opening + closing) / 2
        if not headcount:
            continue
        key = (m_start, m_end)
        if key not in cache_map:
            cache_map[key] = count_working_days(org, m_start, m_end)
        total += cache_map[key] * headcount
    return total


def _exit_date_map(org, user_ids) -> tuple[dict, dict]:
    """user_id -> (last working day, resignation reason) from offboarding records."""
    dates: dict = {}
    reasons: dict = {}
    try:
        from apps.lifecycle.models import OffboardingWorkflow
    except Exception:  # pragma: no cover - lifecycle always installed
        return dates, reasons
    qs = (
        OffboardingWorkflow.objects
        .filter(organization=org)
        .exclude(status=OffboardingWorkflow.Status.CANCELLED)
        .values("user_id", "last_working_day", "resignation_reason")
        .order_by("last_working_day")
    )
    for row in qs:
        dates[row["user_id"]] = row["last_working_day"]
        reasons[row["user_id"]] = row["resignation_reason"] or ""
    return dates, reasons


def _salary_map(org) -> dict:
    """user_id -> active monthly CTC."""
    out: dict = {}
    try:
        from apps.payroll.models import EmployeeSalary
    except Exception:  # pragma: no cover
        return out
    today = timezone.localdate()
    qs = (
        EmployeeSalary.objects
        .filter(user__organization=org, is_active=True, effective_from__lte=today)
        .values("user_id", "monthly_ctc", "effective_from")
        .order_by("effective_from")
    )
    for row in qs:
        out[row["user_id"]] = _f(row["monthly_ctc"])
    return out


def load_workforce(org, filters: HRFilters) -> Workforce:
    """Build the flattened population once per request."""
    dept_map = {
        d["id"]: d["name"]
        for d in Department.objects.filter(organization=org).values("id", "name")
    }

    qs = User.objects.filter(organization=org, role__in=WORKFORCE_ROLES)
    if filters.department:
        try:
            qs = qs.filter(department_id=filters.department)
            qs.exists()
        except (ValueError, ValidationError):
            filters.department = ""
            qs = User.objects.filter(organization=org, role__in=WORKFORCE_ROLES)
    if filters.employment_type:
        qs = qs.filter(employment_type=filters.employment_type)
    if filters.work_mode:
        qs = qs.filter(work_mode=filters.work_mode)
    if filters.location:
        qs = qs.filter(work_location=filters.location)

    people = list(qs.values(
        "id", "first_name", "last_name", "employee_id", "department_id",
        "designation", "gender", "employment_type", "work_mode",
        "employment_status", "work_location", "is_active", "date_of_joining",
        "date_joined", "date_of_birth", "archived_at", "reporting_manager_id",
        "job_grade_id", "org_designation_id",
    ))

    grade_names: dict = {}
    grade_levels: dict = {}
    try:
        from apps.grades.models import Grade

        for g in Grade.objects.filter(organization=org).values("id", "name", "level_number"):
            grade_names[g["id"]] = g["name"]
            grade_levels[g["id"]] = g["level_number"]
    except Exception:  # pragma: no cover
        pass

    exit_dates, exit_reasons = _exit_date_map(org, [p["id"] for p in people])
    salaries = _salary_map(org)

    manager_counts: dict = defaultdict(int)
    for p in people:
        if p["reporting_manager_id"]:
            manager_counts[p["reporting_manager_id"]] += 1

    rows: list[EmployeeRow] = []
    for p in people:
        join_date = p["date_of_joining"]
        if not join_date and p["date_joined"]:
            join_date = timezone.localtime(p["date_joined"]).date()
        exit_date = exit_dates.get(p["id"])
        if not exit_date and not p["is_active"] and p["archived_at"]:
            exit_date = timezone.localtime(p["archived_at"]).date()
        name = " ".join(x for x in (p["first_name"], p["last_name"]) if x).strip()
        rows.append(EmployeeRow(
            id=str(p["id"]),
            name=name or p["employee_id"] or "Unnamed",
            employee_code=p["employee_id"] or "",
            department_id=p["department_id"],
            department=dept_map.get(p["department_id"], "Unassigned"),
            designation=p["designation"] or "",
            grade=grade_names.get(p["job_grade_id"], ""),
            grade_level=grade_levels.get(p["job_grade_id"]),
            gender=p["gender"] or "",
            employment_type=p["employment_type"] or "",
            work_mode=p["work_mode"] or "",
            employment_status=p["employment_status"] or "",
            location=p["work_location"] or "",
            is_active=p["is_active"],
            join_date=join_date,
            exit_date=exit_date,
            exit_reason=exit_reasons.get(p["id"], ""),
            birth_date=p["date_of_birth"],
            monthly_ctc=salaries.get(p["id"], 0.0),
            manager_id=p["reporting_manager_id"],
        ))

    return Workforce(
        org=org,
        filters=filters,
        rows=rows,
        departments=dept_map,
        manager_report_counts=dict(manager_counts),
    )


# -- label helpers ------------------------------------------------------------

def _choice_label(enum_cls, value: str, fallback: str = "Not set") -> str:
    if not value:
        return fallback
    try:
        return enum_cls(value).label
    except ValueError:
        return value.replace("_", " ").title()


def _gender_label(value: str) -> str:
    return _choice_label(User.Gender, value, "Not disclosed")


def _reason_label(value: str) -> str:
    try:
        from apps.lifecycle.models import OffboardingWorkflow

        return _choice_label(OffboardingWorkflow.ResignationReason, value, "Not recorded")
    except Exception:  # pragma: no cover
        return value or "Not recorded"


def _distribution(rows: list[EmployeeRow], key, labeller=None, limit: int = 14) -> dict:
    counter: Counter = Counter()
    for r in rows:
        counter[key(r)] += 1
    items = _top_n(dict(counter), limit)
    return {
        "labels": [(labeller(k) if labeller else (k or "Not set")) for k, _ in items],
        "values": [v for _, v in items],
    }


# -- section: workforce -------------------------------------------------------

def build_workforce(wf: Workforce) -> dict:
    """Headcount, structure, tenure and movement."""
    f = wf.filters
    today = f.date_to
    active = wf.active(today)

    # Headcount / joiner / leaver trend, month by month.
    labels, headcounts, joiners, leavers, net = [], [], [], [], []
    for y, m in f.trend_months:
        m_start = date(y, m, 1)
        m_end = min(_month_end(y, m), today)
        labels.append(_month_label(y, m))
        headcounts.append(wf.headcount(m_end))
        j = len(wf.joiners(m_start, m_end))
        l = len(wf.leavers(m_start, m_end))
        joiners.append(j)
        leavers.append(-l)
        net.append(j - l)

    # Tenure distribution.
    tenure_counter: Counter = Counter()
    tenures: list[float] = []
    for r in active:
        t = r.tenure_years(today)
        if t is None:
            continue
        tenures.append(t)
        band = _band_for(t, TENURE_BANDS)
        if band:
            tenure_counter[band] += 1

    # Department headcount, ordered biggest first.
    dept_counter: Counter = Counter()
    for r in active:
        dept_counter[r.department] += 1
    dept_items = _top_n(dict(dept_counter), 14)

    # Span of control: only count managers who are still active.
    active_ids = {r.id for r in active}
    managers = [
        mid for mid, count in wf.manager_report_counts.items()
        if count > 0 and str(mid) in active_ids
    ]
    span_values = [wf.manager_report_counts[m] for m in managers]

    # Grade pyramid (1 = most senior).
    grade_counter: dict[tuple[int, str], int] = defaultdict(int)
    for r in active:
        if r.grade:
            grade_counter[(r.grade_level or 99, r.grade)] += 1
    grade_items = sorted(grade_counter.items(), key=lambda kv: kv[0][0])

    # Probation / notice pipeline.
    status_counter: Counter = Counter()
    for r in active:
        status_counter[r.employment_status or "ACTIVE"] += 1

    return {
        "trend": {
            "labels": labels,
            "headcount": headcounts,
            "joiners": joiners,
            "leavers": leavers,
            "net": net,
        },
        "tenure": {
            "labels": [b[0] for b in TENURE_BANDS],
            "values": [tenure_counter.get(b[0], 0) for b in TENURE_BANDS],
            "average_years": round(sum(tenures) / len(tenures), 1) if tenures else 0.0,
            "median_years": round(sorted(tenures)[len(tenures) // 2], 1) if tenures else 0.0,
        },
        "departments": {
            "labels": [k for k, _ in dept_items],
            "values": [v for _, v in dept_items],
        },
        "employment_type": _distribution(
            active, lambda r: r.employment_type,
            lambda v: _choice_label(User.EmploymentType, v, "Unspecified"),
        ),
        "work_mode": _distribution(
            active, lambda r: r.work_mode,
            lambda v: _choice_label(User.WorkMode, v, "Unspecified"),
        ),
        "employment_status": {
            "labels": [_choice_label(User.EmploymentStatus, k, "Active") for k, _ in _top_n(dict(status_counter))],
            "values": [v for _, v in _top_n(dict(status_counter))],
        },
        "grades": {
            "labels": [name for (_lvl, name), _ in grade_items],
            "values": [v for _, v in grade_items],
        },
        "locations": _distribution(active, lambda r: r.location or "Unassigned", None, 10),
        "span": {
            "managers": len(managers),
            "average": _div(sum(span_values), len(span_values)),
            "widest": max(span_values) if span_values else 0,
            "individual_contributors": len(active) - len(managers),
        },
        "totals": {
            "headcount": len(active),
            "joiners": len(wf.joiners(f.date_from, f.date_to)),
            "leavers": len(wf.leavers(f.date_from, f.date_to)),
            "on_probation": status_counter.get("PROBATION", 0),
            "on_notice": status_counter.get("NOTICE", 0),
        },
    }


# -- section: attrition -------------------------------------------------------

def build_attrition(wf: Workforce) -> dict:
    """Separation analytics: rate trend, drivers, hot spots, early attrition."""
    f = wf.filters
    today = f.date_to
    leavers = wf.leavers(f.date_from, f.date_to)

    # Monthly attrition-rate trend, split voluntary / involuntary.
    labels, rates, vol, invol = [], [], [], []
    for y, m in f.trend_months:
        m_start = date(y, m, 1)
        m_end = min(_month_end(y, m), today)
        month_leavers = wf.leavers(m_start, m_end)
        avg_hc = wf.average_headcount(m_start, m_end)
        labels.append(_month_label(y, m))
        rates.append(round(len(month_leavers) / avg_hc * 100, 2) if avg_hc else 0.0)
        vol.append(sum(1 for r in month_leavers if r.exit_reason in VOLUNTARY_REASONS))
        invol.append(sum(1 for r in month_leavers if r.exit_reason in INVOLUNTARY_REASONS))

    # Attrition rate by department (leavers / average departmental headcount).
    dept_rows = []
    for dept_id, dept_name in list(wf.departments.items()) + [(None, "Unassigned")]:
        members = [r for r in wf.rows if r.department_id == dept_id]
        if not members:
            continue
        opening = sum(1 for r in members if r.active_on(f.date_from))
        closing = sum(1 for r in members if r.active_on(today))
        avg = (opening + closing) / 2
        dept_leavers = [r for r in members if r.exit_date and f.date_from <= r.exit_date <= today]
        if not avg and not dept_leavers:
            continue
        dept_rows.append({
            "label": dept_name,
            "rate": _pct(len(dept_leavers), avg),
            "leavers": len(dept_leavers),
            "headcount": closing,
        })
    dept_rows.sort(key=lambda d: d["rate"], reverse=True)
    dept_rows = dept_rows[:12]

    # Reasons for leaving.
    reason_counter: Counter = Counter()
    for r in leavers:
        reason_counter[r.exit_reason or ""] += 1
    reason_items = _top_n(dict(reason_counter), 10)

    # Attrition by tenure at exit - surfaces onboarding / early-life failures.
    tenure_counter: Counter = Counter()
    early = 0
    for r in leavers:
        t = _years_between(r.join_date, r.exit_date) if r.exit_date else None
        band = _band_for(t, TENURE_BANDS)
        if band:
            tenure_counter[band] += 1
        if t is not None and t < 1:
            early += 1

    voluntary = sum(1 for r in leavers if r.exit_reason in VOLUNTARY_REASONS)
    involuntary = sum(1 for r in leavers if r.exit_reason in INVOLUNTARY_REASONS)
    neutral = sum(1 for r in leavers if r.exit_reason in NEUTRAL_REASONS)
    unclassified = len(leavers) - voluntary - involuntary - neutral

    period_rate = wf.attrition_rate(f.date_from, f.date_to)
    prev_from, prev_to = f.previous_range
    prev_rate = wf.attrition_rate(prev_from, prev_to)

    # Recent separations table.
    recent = sorted(leavers, key=lambda r: r.exit_date or date.min, reverse=True)[:15]
    recent_rows = [{
        "name": r.name,
        "employee_code": r.employee_code,
        "department": r.department,
        "designation": r.designation,
        "exit_date": r.exit_date.isoformat() if r.exit_date else "",
        "exit_date_display": f"{r.exit_date:%d %b %Y}" if r.exit_date else "",
        "reason": _reason_label(r.exit_reason),
        "tenure": round(_years_between(r.join_date, r.exit_date) or 0, 1) if r.exit_date else 0,
    } for r in recent]

    return {
        "trend": {"labels": labels, "rate": rates, "voluntary": vol, "involuntary": invol},
        "by_department": {
            "labels": [d["label"] for d in dept_rows],
            "values": [d["rate"] for d in dept_rows],
            "rows": dept_rows,
        },
        "reasons": {
            "labels": [_reason_label(k) for k, _ in reason_items],
            "values": [v for _, v in reason_items],
        },
        "by_tenure": {
            "labels": [b[0] for b in TENURE_BANDS],
            "values": [tenure_counter.get(b[0], 0) for b in TENURE_BANDS],
        },
        "split": {
            "labels": ["Voluntary", "Involuntary", "Retirement", "Unclassified"],
            "values": [voluntary, involuntary, neutral, max(0, unclassified)],
        },
        "recent": recent_rows,
        "kpis": {
            "attrition_rate": period_rate,
            "attrition_delta": _delta(period_rate, prev_rate),
            "annualised": wf.attrition_rate(f.date_from, f.date_to, annualised=True),
            "retention_rate": round(max(0.0, 100 - period_rate), 1),
            "separations": len(leavers),
            "voluntary_share": _pct(voluntary, len(leavers)),
            "early_attrition": _pct(early, len(leavers)),
            "avg_exit_tenure": round(
                sum((_years_between(r.join_date, r.exit_date) or 0) for r in leavers) / len(leavers), 1
            ) if leavers else 0.0,
        },
    }


# -- section: attendance & leave ----------------------------------------------

def build_attendance(wf: Workforce) -> dict:
    """Absenteeism, punctuality, working hours and leave consumption."""
    from apps.attendance.models import AttendanceRecord
    from apps.attendance.work_calendar import count_working_days

    f = wf.filters
    org = wf.org
    today = f.date_to
    population = {r.id for r in wf.rows}
    active_now = wf.active(today)

    records = list(
        AttendanceRecord.objects
        .filter(user__organization=org, date__gte=f.date_from, date__lte=f.date_to)
        .values("user_id", "date", "status", "check_in", "check_out", "break_minutes")
    )
    records = [r for r in records if str(r["user_id"]) in population]

    dept_of = {r.id: r.department for r in wf.rows}

    # Shift start times, for the punctuality metric.
    shift_start: dict = {}
    grace: dict = {}
    try:
        from apps.attendance.models import WorkShift

        default_shift = WorkShift.objects.filter(organization=org, is_default=True).first()
        user_shifts = dict(
            User.objects.filter(organization=org, work_shift__isnull=False)
            .values_list("id", "work_shift_id")
        )
        shifts = {s.id: s for s in WorkShift.objects.filter(organization=org)}
        for uid, sid in user_shifts.items():
            s = shifts.get(sid)
            if s:
                shift_start[str(uid)] = s.start_time
                grace[str(uid)] = s.grace_minutes or 0
        if default_shift:
            shift_start.setdefault("__default__", default_shift.start_time)
            grace["__default__"] = default_shift.grace_minutes or 0
    except Exception:  # pragma: no cover
        pass

    status_counter: Counter = Counter()
    by_month: dict = defaultdict(lambda: Counter())
    dept_stats: dict = defaultdict(lambda: {"present": 0, "absent": 0, "late": 0, "total": 0})
    hours_by_month: dict = defaultdict(list)
    late_count = 0
    checkin_count = 0
    overtime_minutes = 0

    for rec in records:
        uid = str(rec["user_id"])
        status = rec["status"]
        key = (rec["date"].year, rec["date"].month)
        status_counter[status] += 1
        by_month[key][status] += 1
        dept = dept_of.get(uid, "Unassigned")
        dept_stats[dept]["total"] += 1
        if status == "PRESENT":
            dept_stats[dept]["present"] += 1
        elif status == "ABSENT":
            dept_stats[dept]["absent"] += 1

        ci, co = rec["check_in"], rec["check_out"]
        if ci:
            checkin_count += 1
            local_in = timezone.localtime(ci) if timezone.is_aware(ci) else ci
            start = shift_start.get(uid) or shift_start.get("__default__")
            allowance = grace.get(uid, grace.get("__default__", 0))
            if start:
                threshold = (
                    datetime.combine(rec["date"], start) + timedelta(minutes=allowance)
                ).time()
                if local_in.time() > threshold:
                    late_count += 1
                    dept_stats[dept]["late"] += 1
        if ci and co:
            local_in = timezone.localtime(ci) if timezone.is_aware(ci) else ci
            local_out = timezone.localtime(co) if timezone.is_aware(co) else co
            mins = int((local_out - local_in).total_seconds() // 60) - (rec["break_minutes"] or 0)
            if 0 < mins < 24 * 60:
                hours_by_month[key].append(mins / 60)
                if mins > 9 * 60:
                    overtime_minutes += mins - 9 * 60

    # Scheduled capacity: working days x headcount, accumulated per month.
    working_days = count_working_days(org, f.date_from, f.date_to)
    wd_cache = wf.cache.setdefault("_working_days", {})
    scheduled = scheduled_capacity(org, wf.rows, f.date_from, f.date_to, wd_cache)

    present = status_counter.get("PRESENT", 0)
    absent = status_counter.get("ABSENT", 0)
    on_leave = status_counter.get("LEAVE", 0)
    half_day = status_counter.get("HALF_DAY", 0)
    wfh = status_counter.get("WFH", 0)

    labels, attendance_pct, absence_pct, avg_hours = [], [], [], []
    for y, m in f.trend_months:
        m_start = date(y, m, 1)
        m_end = min(_month_end(y, m), today)
        counts = by_month.get((y, m), Counter())
        capacity = scheduled_capacity(org, wf.rows, m_start, m_end, wd_cache)
        labels.append(_month_label(y, m))
        attendance_pct.append(_pct(counts.get("PRESENT", 0) + counts.get("WFH", 0), capacity))
        absence_pct.append(_pct(counts.get("ABSENT", 0), capacity))
        vals = hours_by_month.get((y, m), [])
        avg_hours.append(round(sum(vals) / len(vals), 2) if vals else None)

    dept_items = sorted(
        (
            {
                "label": name,
                "rate": _pct(s["present"] + 0, s["total"]),
                "absence": _pct(s["absent"], s["total"]),
                "late": _pct(s["late"], s["total"]),
            }
            for name, s in dept_stats.items() if s["total"]
        ),
        key=lambda d: d["rate"], reverse=True,
    )[:12]

    leave = _leave_metrics(wf, active_now)

    return {
        "trend": {
            "labels": labels,
            "attendance": attendance_pct,
            "absence": absence_pct,
            "avg_hours": avg_hours,
        },
        "status_mix": {
            "labels": ["Present", "Work from home", "Leave", "Half day", "Absent"],
            "values": [present, wfh, on_leave, half_day, absent],
        },
        "by_department": {
            "labels": [d["label"] for d in dept_items],
            "attendance": [d["rate"] for d in dept_items],
            "absence": [d["absence"] for d in dept_items],
            "late": [d["late"] for d in dept_items],
        },
        "leave": leave,
        "kpis": {
            "attendance_rate": _pct(present + wfh, scheduled),
            "absenteeism_rate": _pct(absent, scheduled),
            "punctuality_rate": round(max(0.0, 100 - _pct(late_count, checkin_count)), 1),
            "late_arrivals": late_count,
            "avg_hours": round(
                sum(sum(v) for v in hours_by_month.values())
                / max(1, sum(len(v) for v in hours_by_month.values())), 2
            ),
            "overtime_hours": round(overtime_minutes / 60, 1),
            "working_days": working_days,
            "leave_days": on_leave + half_day * 0.5,
        },
    }


def _leave_metrics(wf: Workforce, active_rows: list[EmployeeRow]) -> dict:
    """Leave utilisation by type plus the accrued liability in currency."""
    empty = {
        "labels": [], "allocated": [], "used": [], "utilisation": [],
        "liability": 0.0, "unused_days": 0.0, "utilisation_rate": 0.0,
        "pending_requests": 0, "approval_rate": 0.0, "avg_approval_hours": 0.0,
    }
    try:
        from apps.leaves.models import LeaveBalance, LeaveRequest, LeaveType
    except Exception:  # pragma: no cover
        return empty

    f = wf.filters
    org = wf.org
    population = {r.id for r in active_rows}
    year = f.date_to.year

    # Utilisation is only meaningful for paid, consumable entitlements. Unpaid
    # buckets like Loss of Pay carry a nominal 365-day ceiling that would swamp
    # every ratio here, so they are excluded.
    types = {
        t["id"]: t
        for t in LeaveType.objects
        .filter(organization=org, is_active=True, is_paid=True)
        .values("id", "name", "carry_forward_max")
    }
    # Only leave that carries forward is a real accrued obligation.
    encashable = {tid for tid, t in types.items() if _f(t["carry_forward_max"]) > 0}

    allocated: dict = defaultdict(float)
    used: dict = defaultdict(float)
    for b in LeaveBalance.objects.filter(
        user__organization=org, year=year, leave_type_id__in=list(types)
    ).values("leave_type_id", "allocated", "used", "adjusted", "carried_forward", "user_id"):
        if str(b["user_id"]) not in population:
            continue
        allocated[b["leave_type_id"]] += _f(b["allocated"]) + _f(b["adjusted"]) + _f(b["carried_forward"])
        used[b["leave_type_id"]] += _f(b["used"])

    labels, alloc_vals, used_vals, util_vals = [], [], [], []
    for tid, meta in types.items():
        a, u = allocated.get(tid, 0.0), used.get(tid, 0.0)
        if a == 0 and u == 0:
            continue
        labels.append(meta["name"])
        alloc_vals.append(round(a, 1))
        used_vals.append(round(u, 1))
        util_vals.append(_pct(u, a))

    total_alloc = sum(alloc_vals)
    total_used = sum(used_vals)
    unused = max(0.0, sum(
        max(0.0, allocated.get(tid, 0.0) - used.get(tid, 0.0)) for tid in encashable
    ))

    # Liability = unused encashable days x average day rate (monthly CTC / 26).
    ctcs = [r.monthly_ctc for r in active_rows if r.monthly_ctc]
    day_rate = (sum(ctcs) / len(ctcs) / 26) if ctcs else 0.0

    requests = list(
        LeaveRequest.objects
        .filter(user__organization=org, start_date__gte=f.date_from, start_date__lte=f.date_to)
        .values("status", "applied_at", "reviewed_at")
    )
    decided = [r for r in requests if r["status"] in ("APPROVED", "REJECTED")]
    approved = [r for r in requests if r["status"] == "APPROVED"]
    turnarounds = [
        (r["reviewed_at"] - r["applied_at"]).total_seconds() / 3600
        for r in decided if r["reviewed_at"] and r["applied_at"]
    ]

    return {
        "labels": labels,
        "allocated": alloc_vals,
        "used": used_vals,
        "utilisation": util_vals,
        "liability": round(unused * day_rate, 2),
        "unused_days": round(unused, 1),
        "utilisation_rate": _pct(total_used, total_alloc),
        "pending_requests": sum(1 for r in requests if r["status"] == "PENDING"),
        "approval_rate": _pct(len(approved), len(decided)),
        "avg_approval_hours": round(sum(turnarounds) / len(turnarounds), 1) if turnarounds else 0.0,
    }


# -- section: compensation ----------------------------------------------------

SALARY_BANDS = (
    ("< 25K", 0, 25_000),
    ("25K-50K", 25_000, 50_000),
    ("50K-75K", 50_000, 75_000),
    ("75K-1L", 75_000, 100_000),
    ("1L-1.5L", 100_000, 150_000),
    ("1.5L-2.5L", 150_000, 250_000),
    ("2.5L+", 250_000, 10**12),
)


def build_compensation(wf: Workforce) -> dict:
    """Payroll cost, salary distribution, pay equity and cost efficiency."""
    f = wf.filters
    today = f.date_to
    active = wf.active(today)

    labels, gross, net, deductions, cpe, headcounts = [], [], [], [], [], []
    run_status: Counter = Counter()
    payment_status: Counter = Counter()
    latest_period = ""
    total_gross = total_net = total_deductions = total_employer = 0.0
    total_overtime = total_bonus = 0.0
    processed_months: set = set()

    try:
        from apps.payroll.models import PayrollRun, Payslip

        runs = {
            (r["year"], r["month"]): r
            for r in PayrollRun.objects.filter(organization=wf.org).values(
                "id", "year", "month", "status", "total_gross", "total_net",
                "total_deductions", "total_bonus", "employee_count",
            )
        }
        for y, m in f.trend_months:
            run = runs.get((y, m))
            labels.append(_month_label(y, m))
            if not run:
                gross.append(None); net.append(None); deductions.append(None)
                cpe.append(None); headcounts.append(None)
                continue
            g = _f(run["total_gross"]); n = _f(run["total_net"]); d = _f(run["total_deductions"])
            count = run["employee_count"] or 0
            gross.append(round(g, 2))
            net.append(round(n, 2))
            deductions.append(round(d, 2))
            cpe.append(round(g / count, 2) if count else None)
            headcounts.append(count)

        period_months = set(f.months)
        for (y, m), run in runs.items():
            if (y, m) in period_months:
                processed_months.add((y, m))
                run_status[run["status"]] += 1
                total_gross += _f(run["total_gross"])
                total_net += _f(run["total_net"])
                total_deductions += _f(run["total_deductions"])
                total_bonus += _f(run["total_bonus"])

        slips = list(
            Payslip.objects
            .filter(payroll_run__organization=wf.org,
                    payroll_run__year__gte=f.date_from.year)
            .values("payroll_run__year", "payroll_run__month", "payment_status",
                    "employer_pf", "overtime_amount")
        )
        month_set = set(f.months)
        for s in slips:
            if (s["payroll_run__year"], s["payroll_run__month"]) not in month_set:
                continue
            payment_status[s["payment_status"]] += 1
            total_employer += _f(s["employer_pf"])
            total_overtime += _f(s["overtime_amount"])

        if runs:
            ly, lm = max(runs.keys())
            latest_period = f"{calendar.month_name[lm]} {ly}"
    except Exception:  # pragma: no cover - payroll optional
        pass

    # Salary band distribution from live CTC records.
    band_counter: Counter = Counter()
    ctcs = [r.monthly_ctc for r in active if r.monthly_ctc]
    for value in ctcs:
        for label, low, high in SALARY_BANDS:
            if low <= value < high:
                band_counter[label] += 1
                break

    # Average CTC by department.
    dept_ctc: dict = defaultdict(list)
    for r in active:
        if r.monthly_ctc:
            dept_ctc[r.department].append(r.monthly_ctc)
    dept_items = sorted(
        ((name, sum(v) / len(v)) for name, v in dept_ctc.items()),
        key=lambda kv: kv[1], reverse=True,
    )[:12]

    # Average CTC by grade level (compensation pyramid).
    grade_ctc: dict = defaultdict(list)
    for r in active:
        if r.monthly_ctc and r.grade:
            grade_ctc[(r.grade_level or 99, r.grade)].append(r.monthly_ctc)
    grade_items = sorted(grade_ctc.items(), key=lambda kv: kv[0][0])

    sorted_ctcs = sorted(ctcs)
    median_ctc = sorted_ctcs[len(sorted_ctcs) // 2] if sorted_ctcs else 0.0
    mean_ctc = sum(ctcs) / len(ctcs) if ctcs else 0.0
    # Divide only by months that actually have a processed run in the period -
    # the trend window can be wider than the selected range.
    payroll_months = max(1, len(processed_months))

    return {
        "trend": {
            "labels": labels, "gross": gross, "net": net,
            "deductions": deductions, "cost_per_employee": cpe, "headcount": headcounts,
        },
        "bands": {
            "labels": [b[0] for b in SALARY_BANDS],
            "values": [band_counter.get(b[0], 0) for b in SALARY_BANDS],
        },
        "by_department": {
            "labels": [k for k, _ in dept_items],
            "values": [round(v, 2) for _, v in dept_items],
        },
        "by_grade": {
            "labels": [name for (_lvl, name), _ in grade_items],
            "values": [round(sum(v) / len(v), 2) for _, v in grade_items],
        },
        "run_status": {
            "labels": [k.replace("_", " ").title() for k in run_status],
            "values": list(run_status.values()),
        },
        "payment_status": {
            "labels": [k.replace("_", " ").title() for k in payment_status],
            "values": list(payment_status.values()),
            "period": latest_period,
        },
        "kpis": {
            "total_gross": round(total_gross, 2),
            "total_net": round(total_net, 2),
            "total_deductions": round(total_deductions, 2),
            "employer_contribution": round(total_employer, 2),
            "total_cost": round(total_gross + total_employer, 2),
            "monthly_run_rate": round((total_gross + total_employer) / payroll_months, 2),
            "cost_per_employee": round(total_gross / max(1, len(active)) / payroll_months, 2),
            "average_ctc": round(mean_ctc, 2),
            "median_ctc": round(median_ctc, 2),
            "overtime_cost": round(total_overtime, 2),
            "bonus": round(total_bonus, 2),
            "payroll_months": payroll_months,
            "annualised_cost": round((total_gross + total_employer) / payroll_months * 12, 2),
        },
    }


# -- section: diversity -------------------------------------------------------

def build_diversity(wf: Workforce) -> dict:
    """Representation, age profile and pay equity."""
    f = wf.filters
    today = f.date_to
    active = wf.active(today)

    gender_counter: Counter = Counter()
    for r in active:
        gender_counter[r.gender or ""] += 1
    gender_items = _top_n(dict(gender_counter), 6)

    # Representation by department.
    dept_gender: dict = defaultdict(lambda: Counter())
    for r in active:
        dept_gender[r.department][r.gender or ""] += 1
    dept_names = [name for name, _ in _top_n({k: sum(v.values()) for k, v in dept_gender.items()}, 12)]
    female_by_dept, male_by_dept, other_by_dept = [], [], []
    for name in dept_names:
        c = dept_gender[name]
        female_by_dept.append(c.get("FEMALE", 0))
        male_by_dept.append(c.get("MALE", 0))
        other_by_dept.append(sum(v for k, v in c.items() if k not in ("MALE", "FEMALE")))

    # Age profile.
    age_counter: Counter = Counter()
    ages = []
    for r in active:
        a = r.age(today)
        if a is None:
            continue
        ages.append(a)
        band = _band_for(a, AGE_BANDS)
        if band:
            age_counter[band] += 1

    # Leadership = senior grades (1-3) plus anyone with direct reports.
    manager_ids = {str(m) for m, count in wf.manager_report_counts.items() if count}
    leaders = [
        r for r in active
        if (r.grade_level is not None and r.grade_level <= 3) or r.id in manager_ids
    ]
    female_leaders = sum(1 for r in leaders if r.gender == "FEMALE")

    # Pay equity.
    male_pay = [r.monthly_ctc for r in active if r.gender == "MALE" and r.monthly_ctc]
    female_pay = [r.monthly_ctc for r in active if r.gender == "FEMALE" and r.monthly_ctc]
    male_mean = sum(male_pay) / len(male_pay) if male_pay else 0.0
    female_mean = sum(female_pay) / len(female_pay) if female_pay else 0.0
    # Undefined rather than 100% when one cohort has no salary data at all.
    pay_gap = _pct(male_mean - female_mean, male_mean) if (male_pay and female_pay) else None

    # Pay gap by department, where both genders are represented.
    gap_labels, gap_values = [], []
    for name in dept_names:
        m = [r.monthly_ctc for r in active if r.department == name and r.gender == "MALE" and r.monthly_ctc]
        fem = [r.monthly_ctc for r in active if r.department == name and r.gender == "FEMALE" and r.monthly_ctc]
        if not m or not fem:
            continue
        m_avg = sum(m) / len(m)
        gap_labels.append(name)
        gap_values.append(_pct(m_avg - sum(fem) / len(fem), m_avg))

    # Hiring diversity - gender split of joiners over the trend window.
    hire_labels, hire_female, hire_male = [], [], []
    for y, m in f.trend_months:
        m_start = date(y, m, 1)
        m_end = min(_month_end(y, m), today)
        joiners = wf.joiners(m_start, m_end)
        hire_labels.append(_month_label(y, m))
        hire_female.append(sum(1 for r in joiners if r.gender == "FEMALE"))
        hire_male.append(sum(1 for r in joiners if r.gender == "MALE"))

    total = len(active)
    return {
        "gender": {
            "labels": [_gender_label(k) for k, _ in gender_items],
            "values": [v for _, v in gender_items],
        },
        "gender_by_department": {
            "labels": dept_names,
            "female": female_by_dept,
            "male": male_by_dept,
            "other": other_by_dept,
        },
        "age": {
            "labels": [b[0] for b in AGE_BANDS],
            "values": [age_counter.get(b[0], 0) for b in AGE_BANDS],
        },
        "employment_type": _distribution(
            active, lambda r: r.employment_type,
            lambda v: _choice_label(User.EmploymentType, v, "Unspecified"),
        ),
        "pay_gap_by_department": {"labels": gap_labels, "values": gap_values},
        "hiring": {"labels": hire_labels, "female": hire_female, "male": hire_male},
        "kpis": {
            "female_share": _pct(gender_counter.get("FEMALE", 0), total),
            "male_share": _pct(gender_counter.get("MALE", 0), total),
            "leadership_female_share": _pct(female_leaders, len(leaders)),
            "leaders": len(leaders),
            "pay_gap": pay_gap,
            "average_age": round(sum(ages) / len(ages), 1) if ages else 0.0,
            "under_35_share": _pct(
                age_counter.get("Under 25", 0) + age_counter.get("25-34", 0), sum(age_counter.values())
            ),
            "disclosure_rate": _pct(total - gender_counter.get("", 0), total),
        },
    }


# -- section: department scorecard --------------------------------------------

def build_scorecard(wf: Workforce) -> dict:
    """One row per department, every headline metric side by side."""
    from apps.attendance.models import AttendanceRecord
    from apps.attendance.work_calendar import count_working_days

    f = wf.filters
    today = f.date_to
    population = {r.id: r for r in wf.rows}

    att: dict = defaultdict(lambda: {"present": 0, "absent": 0, "total": 0})
    records = (
        AttendanceRecord.objects
        .filter(user__organization=wf.org, date__gte=f.date_from, date__lte=f.date_to)
        .values("user_id", "status")
    )
    for rec in records:
        row = population.get(str(rec["user_id"]))
        if not row:
            continue
        bucket = att[row.department]
        bucket["total"] += 1
        if rec["status"] in ("PRESENT", "WFH"):
            bucket["present"] += 1
        elif rec["status"] == "ABSENT":
            bucket["absent"] += 1

    working_days = count_working_days(wf.org, f.date_from, f.date_to)
    wd_cache = wf.cache.setdefault("_working_days", {})

    rows = []
    dept_names = sorted({r.department for r in wf.rows})
    for name in dept_names:
        members = [r for r in wf.rows if r.department == name]
        active = [r for r in members if r.active_on(today)]
        opening = sum(1 for r in members if r.active_on(f.date_from))
        joiners = [r for r in members if r.join_date and f.date_from <= r.join_date <= today]
        leavers = [r for r in members if r.exit_date and f.date_from <= r.exit_date <= today]
        avg_hc = (opening + len(active)) / 2
        capacity = scheduled_capacity(wf.org, members, f.date_from, today, wd_cache)
        tenures = [t for t in (r.tenure_years(today) for r in active) if t is not None]
        ctcs = [r.monthly_ctc for r in active if r.monthly_ctc]
        a = att.get(name, {"present": 0, "absent": 0, "total": 0})
        rows.append({
            "department": name,
            "headcount": len(active),
            "opening": opening,
            "joiners": len(joiners),
            "leavers": len(leavers),
            "net_change": len(joiners) - len(leavers),
            "attrition_rate": _pct(len(leavers), avg_hc),
            "avg_tenure": round(sum(tenures) / len(tenures), 1) if tenures else 0.0,
            "attendance_rate": _pct(a["present"], capacity),
            "absenteeism_rate": _pct(a["absent"], capacity),
            "avg_ctc": round(sum(ctcs) / len(ctcs), 0) if ctcs else 0,
            "monthly_cost": round(sum(ctcs), 0) if ctcs else 0,
            "female_share": _pct(sum(1 for r in active if r.gender == "FEMALE"), len(active)),
            "on_notice": sum(1 for r in active if r.employment_status == "NOTICE"),
        })

    rows.sort(key=lambda r: r["headcount"], reverse=True)

    totals = {
        "headcount": sum(r["headcount"] for r in rows),
        "joiners": sum(r["joiners"] for r in rows),
        "leavers": sum(r["leavers"] for r in rows),
        "net_change": sum(r["net_change"] for r in rows),
        "monthly_cost": sum(r["monthly_cost"] for r in rows),
        "on_notice": sum(r["on_notice"] for r in rows),
    }
    return {"rows": rows, "totals": totals, "working_days": working_days}


# -- section: executive overview ----------------------------------------------

def build_overview(wf: Workforce) -> dict:
    """The headline KPI band plus the two charts an exec actually looks at."""
    f = wf.filters
    today = f.date_to
    prev_from, prev_to = f.previous_range
    active = wf.active(today)

    headcount = len(active)
    prev_headcount = wf.headcount(prev_to)
    joiners = wf.joiners(f.date_from, f.date_to)
    leavers = wf.leavers(f.date_from, f.date_to)
    prev_joiners = wf.joiners(prev_from, prev_to)
    prev_leavers = wf.leavers(prev_from, prev_to)

    attrition = wf.attrition_rate(f.date_from, f.date_to)
    prev_attrition = wf.attrition_rate(prev_from, prev_to)

    tenures = [t for t in (r.tenure_years(today) for r in active) if t is not None]
    avg_tenure = round(sum(tenures) / len(tenures), 1) if tenures else 0.0

    symbol = currency_symbol(wf.org)
    attendance = section_data(wf, "attendance")
    compensation = section_data(wf, "compensation")
    diversity = section_data(wf, "diversity")
    workforce = section_data(wf, "workforce")

    active_ids = {r.id for r in active}
    manager_count = sum(
        1 for mid, count in wf.manager_report_counts.items()
        if count and str(mid) in active_ids
    )

    kpis = [
        {
            "key": "headcount", "label": "Active headcount", "value": headcount,
            "format": "int", "delta": _delta(headcount, prev_headcount),
            "hint": f"{len(joiners)} joined / {len(leavers)} exited this period",
            "icon": "users", "tone": "violet",
        },
        {
            "key": "attrition", "label": "Attrition rate", "value": attrition,
            "format": "pct", "delta": _delta(attrition, prev_attrition), "invert": True,
            "hint": f"{wf.attrition_rate(f.date_from, f.date_to, annualised=True)}% annualised",
            "icon": "user-minus", "tone": "rose",
        },
        {
            "key": "retention", "label": "Retention rate", "value": round(max(0.0, 100 - attrition), 1),
            "format": "pct", "delta": None,
            "hint": f"Avg tenure {avg_tenure} yrs", "icon": "shield-check", "tone": "emerald",
        },
        {
            "key": "hires", "label": "New hires", "value": len(joiners),
            "format": "int", "delta": _delta(len(joiners), len(prev_joiners)),
            "hint": f"Net change {len(joiners) - len(leavers):+d}", "icon": "user-plus", "tone": "sky",
        },
        {
            "key": "absenteeism", "label": "Absenteeism", "value": attendance["kpis"]["absenteeism_rate"],
            "format": "pct", "delta": None, "invert": True,
            "hint": f"Attendance {attendance['kpis']['attendance_rate']}%",
            "icon": "calendar-x", "tone": "amber",
        },
        {
            "key": "cost", "label": "Payroll run rate", "value": compensation["kpis"]["monthly_run_rate"],
            "format": "currency", "delta": None,
            "hint": f"{symbol}{compensation['kpis']['cost_per_employee']:,.0f} per employee / month",
            "icon": "wallet", "tone": "indigo",
        },
        {
            "key": "diversity", "label": "Female representation", "value": diversity["kpis"]["female_share"],
            "format": "pct", "delta": None,
            "hint": f"{diversity['kpis']['leadership_female_share']}% in leadership",
            "icon": "scale", "tone": "pink",
        },
        {
            "key": "span", "label": "Span of control", "value": workforce["span"]["average"],
            "format": "ratio", "delta": None,
            "hint": f"{manager_count} people manager{'' if manager_count == 1 else 's'}",
            "icon": "network", "tone": "teal",
        },
    ]

    return {
        "kpis": kpis,
        "headcount_trend": workforce["trend"],
        "attrition_trend": section_data(wf, "attrition")["trend"],
        "department_mix": workforce["departments"],
        "tenure": workforce["tenure"],
        "leave": attendance["leave"],
        "insights": build_insights(wf, workforce, attendance, compensation, diversity),
        "totals": {
            "headcount": headcount,
            "joiners": len(joiners),
            "leavers": len(leavers),
            "on_notice": workforce["totals"]["on_notice"],
            "on_probation": workforce["totals"]["on_probation"],
            "departments": len({r.department for r in active}),
            "managers": manager_count,
        },
    }


# -- narrative insights -------------------------------------------------------

def build_insights(wf, workforce: dict, attendance: dict, compensation: dict, diversity: dict) -> list[dict]:
    """Plain-language findings, ranked by how much they should worry HR."""
    f = wf.filters
    out: list[dict] = []
    attrition = section_data(wf, "attrition")
    org_rate = attrition["kpis"]["attrition_rate"]

    # Departments running hot on attrition.
    for row in attrition["by_department"]["rows"][:3]:
        if row["leavers"] >= 2 and org_rate and row["rate"] > org_rate * 1.5:
            out.append({
                "tone": "critical", "icon": "trending-up",
                "title": f"{row['label']} attrition is {round(row['rate'] / org_rate, 1)}x the org average",
                "body": f"{row['leavers']} exits against a {row['headcount']}-person team "
                        f"({row['rate']}% vs {org_rate}% org-wide). Worth an exit-interview review.",
            })

    early = attrition["kpis"]["early_attrition"]
    if early >= 30 and attrition["kpis"]["separations"] >= 3:
        out.append({
            "tone": "warning", "icon": "user-x",
            "title": f"{early}% of exits happened within the first year",
            "body": "High early attrition usually points at hiring fit or onboarding, "
                    "not compensation. Check the 90-day check-in completion rate.",
        })

    if workforce["totals"]["on_notice"]:
        out.append({
            "tone": "warning", "icon": "clock",
            "title": f"{workforce['totals']['on_notice']} employees are serving notice",
            "body": "Pipeline these roles now - replacement hiring typically runs "
                    "45-60 days behind the last working day.",
        })

    absent = attendance["kpis"]["absenteeism_rate"]
    if absent > 5:
        out.append({
            "tone": "warning", "icon": "calendar-x",
            "title": f"Absenteeism is {absent}%, above the 5% healthy ceiling",
            "body": f"{attendance['kpis']['late_arrivals']} late arrivals in the same window. "
                    "Check shift design and approval turnaround before escalating.",
        })

    punctuality = attendance["kpis"]["punctuality_rate"]
    if punctuality and punctuality < 85:
        out.append({
            "tone": "info", "icon": "alarm-clock",
            "title": f"Punctuality is {punctuality}%",
            "body": "More than one in seven check-ins is late against the rostered shift "
                    "start plus grace period.",
        })

    gap = diversity["kpis"]["pay_gap"]
    if gap is not None and abs(gap) >= 10:
        direction = "higher" if gap > 0 else "lower"
        out.append({
            "tone": "warning" if gap > 0 else "info", "icon": "scale",
            "title": f"Mean pay gap of {abs(gap)}% - male average CTC is {direction}",
            "body": "Compare like-for-like within grades before acting; a raw gap is "
                    "usually a seniority-mix effect, not a same-role gap.",
        })

    leadership = diversity["kpis"]["leadership_female_share"]
    overall_female = diversity["kpis"]["female_share"]
    if overall_female and leadership + 10 < overall_female:
        out.append({
            "tone": "info", "icon": "users",
            "title": f"Female representation drops from {overall_female}% overall to {leadership}% in leadership",
            "body": "A widening gap at senior levels is the classic signal for a "
                    "promotion-pipeline review.",
        })

    liability = attendance["leave"].get("liability", 0)
    if liability > 0:
        out.append({
            "tone": "info", "icon": "palmtree",
            "title": f"Accrued leave liability is {currency_symbol(wf.org)}{liability:,.0f}",
            "body": f"{attendance['leave']['unused_days']} unused days at "
                    f"{attendance['leave']['utilisation_rate']}% utilisation. "
                    "Encourage consumption before the carry-forward cut-off.",
        })

    growth = workforce["totals"]["joiners"] - workforce["totals"]["leavers"]
    if growth > 0:
        out.append({
            "tone": "positive", "icon": "trending-up",
            "title": f"Headcount grew by {growth} over {f.label().lower()}",
            "body": f"{workforce['totals']['joiners']} hires against "
                    f"{workforce['totals']['leavers']} exits across "
                    f"{len(workforce['departments']['labels'])} departments.",
        })

    if not out:
        out.append({
            "tone": "positive", "icon": "check-circle",
            "title": "No material risks flagged for this period",
            "body": "Attrition, absenteeism and pay equity are all inside normal ranges "
                    "for the selected window.",
        })
    return out[:6]


# -- dispatcher & caching -----------------------------------------------------

SECTION_BUILDERS = {
    "overview": build_overview,
    "workforce": build_workforce,
    "attrition": build_attrition,
    "attendance": build_attendance,
    "compensation": build_compensation,
    "diversity": build_diversity,
    "scorecard": build_scorecard,
}


def section_data(wf: Workforce, section: str) -> dict:
    """Build one section, memoised per Workforce so the overview stays cheap."""
    builder = SECTION_BUILDERS.get(section)
    if not builder:
        raise KeyError(section)
    if section not in wf.cache:
        wf.cache[section] = builder(wf)
    return wf.cache[section]


def get_section(org, filters: HRFilters, section: str, use_cache: bool = True) -> dict:
    """Public entry point: cached, filter-scoped payload for one dashboard tab."""
    key = f"hr_analytics:{getattr(org, 'pk', 'na')}:{section}:{filters.cache_suffix()}"
    if use_cache:
        cached = cache.get(key)
        if cached is not None:
            return cached
    payload = section_data(load_workforce(org, filters), section)
    payload = {**payload, "filters": filters.as_dict()}
    if use_cache:
        cache.set(key, payload, CACHE_TTL)
    return payload


def filter_options(org) -> dict:
    """Choices for the filter bar, restricted to values the org actually uses."""
    used_types = set(
        User.objects.filter(organization=org, role__in=WORKFORCE_ROLES)
        .exclude(employment_type="")
        .values_list("employment_type", flat=True).distinct()
    )
    used_modes = set(
        User.objects.filter(organization=org, role__in=WORKFORCE_ROLES)
        .exclude(work_mode="")
        .values_list("work_mode", flat=True).distinct()
    )
    locations = sorted(
        v for v in User.objects.filter(organization=org, role__in=WORKFORCE_ROLES)
        .exclude(work_location="")
        .values_list("work_location", flat=True).distinct()
    )
    return {
        "departments": list(
            Department.objects.filter(organization=org, is_active=True)
            .order_by("name").values("id", "name")
        ),
        "employment_types": [
            {"value": v, "label": User.EmploymentType(v).label}
            for v in User.EmploymentType.values if v in used_types
        ],
        "work_modes": [
            {"value": v, "label": User.WorkMode(v).label}
            for v in User.WorkMode.values if v in used_modes
        ],
        "locations": [{"value": v, "label": v} for v in locations],
        "periods": [{"value": v, "label": label} for v, label in PERIOD_CHOICES],
    }


def invalidate(org) -> None:
    """Drop this org's cached sections (call after a bulk staff/payroll change)."""
    try:
        cache.delete_pattern(f"hr_analytics:{getattr(org, 'pk', 'na')}:*")  # redis
    except AttributeError:
        cache.clear()


# -- exports ------------------------------------------------------------------

SCORECARD_HEADERS = [
    "Department", "Opening headcount", "Closing headcount", "Joiners", "Exits",
    "Net change", "Attrition %", "Avg tenure (yrs)", "Attendance %",
    "Absenteeism %", "Avg monthly CTC", "Monthly salary cost", "Female %", "On notice",
]


def _scorecard_values(row: dict) -> list:
    return [
        row["department"], row["opening"], row["headcount"], row["joiners"],
        row["leavers"], row["net_change"], row["attrition_rate"], row["avg_tenure"],
        row["attendance_rate"], row["absenteeism_rate"], row["avg_ctc"],
        row["monthly_cost"], row["female_share"], row["on_notice"],
    ]


def export_scorecard_csv(org, filters: HRFilters) -> HttpResponse:
    data = get_section(org, filters, "scorecard", use_cache=False)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([f"HR Analytics - department scorecard - {filters.label()}"])
    writer.writerow([f"{filters.date_from:%d %b %Y} to {filters.date_to:%d %b %Y}",
                     f"{data['working_days']} working days"])
    writer.writerow([])
    writer.writerow(SCORECARD_HEADERS)
    for row in data["rows"]:
        writer.writerow(_scorecard_values(row))
    totals = data["totals"]
    writer.writerow([])
    writer.writerow(["TOTAL", "", totals["headcount"], totals["joiners"], totals["leavers"],
                     totals["net_change"], "", "", "", "", "", totals["monthly_cost"], "",
                     totals["on_notice"]])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="hr-analytics-scorecard-{filters.date_to:%Y-%m-%d}.csv"'
    )
    return response


def export_scorecard_xlsx(org, filters: HRFilters) -> HttpResponse:
    """Multi-sheet workbook: scorecard, headcount movement and attrition drivers."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wf = load_workforce(org, filters)
    scorecard = section_data(wf, "scorecard")
    workforce = section_data(wf, "workforce")
    attrition = section_data(wf, "attrition")

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6D28D9")
    title_font = Font(bold=True, size=13)

    def style_header(ws, row_idx: int, width: int) -> None:
        for col in range(1, width + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 1 - department scorecard
    ws = wb.active
    ws.title = "Scorecard"
    ws.append([f"HR Analytics - Department Scorecard"])
    ws["A1"].font = title_font
    ws.append([f"{filters.label()} ({filters.date_from:%d %b %Y} - {filters.date_to:%d %b %Y})"])
    ws.append([])
    ws.append(SCORECARD_HEADERS)
    style_header(ws, 4, len(SCORECARD_HEADERS))
    for row in scorecard["rows"]:
        ws.append(_scorecard_values(row))
    for idx, width in enumerate([22] + [16] * (len(SCORECARD_HEADERS) - 1), start=1):
        ws.column_dimensions[ws.cell(row=4, column=idx).column_letter].width = width
    ws.freeze_panes = "A5"

    # Sheet 2 - headcount movement
    ws2 = wb.create_sheet("Headcount movement")
    ws2.append(["Month", "Closing headcount", "Joiners", "Exits", "Net change"])
    style_header(ws2, 1, 5)
    trend = workforce["trend"]
    for i, label in enumerate(trend["labels"]):
        ws2.append([label, trend["headcount"][i], trend["joiners"][i],
                    abs(trend["leavers"][i]), trend["net"][i]])
    for col in "ABCDE":
        ws2.column_dimensions[col].width = 18
    ws2.freeze_panes = "A2"

    # Sheet 3 - attrition detail
    ws3 = wb.create_sheet("Separations")
    ws3.append(["Employee", "Employee ID", "Department", "Designation",
                "Last working day", "Reason", "Tenure (yrs)"])
    style_header(ws3, 1, 7)
    for row in attrition["recent"]:
        ws3.append([row["name"], row["employee_code"], row["department"],
                    row["designation"], row["exit_date_display"], row["reason"], row["tenure"]])
    for col in "ABCDEFG":
        ws3.column_dimensions[col].width = 20
    ws3.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="hr-analytics-{filters.date_to:%Y-%m-%d}.xlsx"'
    )
    return response
